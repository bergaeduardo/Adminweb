import pprint
import logging
import os
from django.conf import settings
from django import template
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.http import HttpResponse, HttpResponseRedirect,JsonResponse
from django.template import loader
from django.urls import reverse
from django.shortcuts import render, redirect,get_object_or_404
from django.views.decorators.http import require_http_methods
# from Transportes.models import Transporte
# from Transportes.forms import TransporteForm
from django.views import View
from django.views.generic.list import ListView
from apps.settingsUrls import *
from consultasTango.forms import TurnoForm,TurnoEditForm,CodigoErrorForm,CategoriaForm, SubcategoriaForm, RelacionForm, TurnoReservaForm, EstadoTurnoForm
from consultasTango.models import Turno,CodigosError, TurnoReserva, EstadoTurno, HistorialEstadoTurno, IncidenciasTurno
from django.utils import timezone
from django.views.decorators.csrf import csrf_exempt
from django.utils.decorators import method_decorator
from django.core.files.storage import FileSystemStorage
from django.core.paginator import Paginator
from django.contrib.auth.decorators import user_passes_test
import openpyxl
import pandas as pd
import json
from apps.home.SQL.Sql_Tango import * # Keep SQL imports
from apps.home.SQL.Sql_WMS import * # Keep SQL imports
import xlwt # Keep excel writing imports


def formatear_orden_compra(orden_compra):
    """
    Formatea el campo orden_compra para mostrar de forma legible.
    Convierte formatos como "['OC1', 'OC2']" o "OC1|OC2" en "OC1, OC2"
    """
    if not orden_compra:
        return ''
    
    oc_str = str(orden_compra)
    
    # Remover corchetes y comillas de formato lista Python
    oc_str = oc_str.replace('[', '').replace(']', '').replace("'", '').replace('"', '')
    
    # Reemplazar pipes por comas
    oc_str = oc_str.replace('|', ', ')
    
    # Limpiar espacios extras
    oc_str = ', '.join([oc.strip() for oc in oc_str.split(',') if oc.strip()])
    
    return oc_str
import xlrd # Keep excel reading imports
from numpy import int64, isnan # Keep numpy imports
from datetime import datetime, timedelta, date, time as dt_time


# @login_required(login_url="/login/")
# def (request):
#     Nombre=''
#     dir_iframe = DIR_HERAMIENTAS['']
#     return render(request,'home/PlantillaHerramientas.html',{'dir_iframe':dir_iframe,'Nombre':Nombre})


# Logistica

@login_required(login_url="/login/")
def AnularRemitos(request):
    Nombre = 'Anular Remitos'
    dir_iframe = DIR_HERAMIENTAS['AnularRemitos']
    return render(request, 'home/PlantillaHerramientas.html', {'dir_iframe': dir_iframe,'Nombre':Nombre})

@login_required(login_url="/login/")
def RemisionMasiva(request):
    Nombre = ''
    dir_iframe = DIR_HERAMIENTAS['RemisionMasiva']
    return render(request, 'home/PlantillaHerramientas.html', {'dir_iframe': dir_iframe,'Nombre':Nombre})

@login_required(login_url="/login/")
def ImprimirEtiquetasBultos(request):
    Nombre = 'Imprimir Etiquetas Bultos'
    dir_iframe = DIR_HERAMIENTAS['ImprimirEtiquetasBultos']
    return render(request, 'home/PlantillaHerramientas.html', {'dir_iframe': dir_iframe,'Nombre':Nombre})

@login_required(login_url="/login/")
def CargaAnticipoGrupo(request):
    Nombre = 'Carga de anticipos'
    dir_iframe = DIR_HERAMIENTAS['CargaAnticipoGrupo']
    return render(request, 'home/PlantillaHerramientas.html', {'dir_iframe': dir_iframe,'Nombre':Nombre})

@login_required(login_url="/login/")
def registro_turno(request):
    if request.method == 'POST':
        form = TurnoForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect('herramientas:herramientas_listado_turnos') # Updated redirect
    else:
        form = TurnoForm()

    return render(request, 'appConsultasTango/registro_turno.html', {'form': form, 'Nombre': 'Registro de Turno'})

@login_required(login_url="/login/")
def get_nombre_proveedor(request):
    """
    API que retorna nombre de proveedor consultando tabla CPA01 de Tango
    Consulta dinámicamente la base de datos en lugar de usar diccionario estático
    """
    codigo_proveedor = request.GET.get('codigo', '').strip().upper()
    
    if not codigo_proveedor:
        return JsonResponse({'nombre': '', 'error': 'Código no proporcionado'})
    
    try:
        from apps.home.SQL.Sql_Tango import obtener_nombre_proveedor_por_codigo
        nombre = obtener_nombre_proveedor_por_codigo(codigo_proveedor)
        
        if nombre:
            return JsonResponse({'nombre': nombre})
        else:
            return JsonResponse({'nombre': '', 'error': 'Proveedor no encontrado'})
    except Exception as e:
        return JsonResponse({'nombre': '', 'error': f'Error al consultar proveedor: {str(e)}'})


@login_required(login_url="/login/")
def get_ordenes_compra_proveedor(request):
    """
    API AJAX que retorna órdenes de compra activas de un proveedor
    Formato Select2 compatible para cargar dinámicamente en campo multi-selección
    
    Returns JSON con formato:
    {
        "ordenes": [
            {"id": " 0000100012634", "text": " 0000100012634 - Emitida (01/12/2024)"},
            ...
        ]
    }
    """
    codigo_proveedor = request.GET.get('codigo', '').strip().upper()
    
    if not codigo_proveedor:
        return JsonResponse({'ordenes': [], 'error': 'Código no proporcionado'})
    
    try:
        from apps.home.SQL.Sql_Tango import obtener_ordenes_compra_activas_proveedor
        ordenes = obtener_ordenes_compra_activas_proveedor(codigo_proveedor)
        
        # Formatear para Select2
        ordenes_lista = [
            {
                'id': orden[0].strip(),  # N_ORDEN_CO (string con espacio inicial)
                'text': f"{orden[0].strip()} - {orden[2]} ({orden[1].strftime('%d/%m/%Y')})"
            }
            for orden in ordenes
        ]
        
        return JsonResponse({'ordenes': ordenes_lista})
    except Exception as e:
        return JsonResponse({'ordenes': [], 'error': f'Error al consultar órdenes: {str(e)}'})

@login_required(login_url="/login/")
def listado_turnos(request):
    turnos = Turno.objects.all().order_by('-FechaAsignacion')
    for turno in turnos:
        turno.progreso = calcular_progreso(turno)
        turno.save()
    nombre_template = 'Listado de Turnos'
    return render(request, 'appConsultasTango/listado_turnos.html', {'turnos': turnos,'Nombre':nombre_template})

def eliminar_turno(request, turno_id):
    turno = get_object_or_404(Turno, pk=turno_id)
    if request.method == 'POST':
        turno.delete()
        return JsonResponse({'success': True})
    return JsonResponse({'success': False})

def calcular_progreso(turno):
    total_steps = 3
    completed_steps = sum([turno.Recepcionado, turno.Auditado, turno.Posicionado])
    return int((completed_steps / total_steps) * 100)

def ver_turno(request, turno_id):
    turno = get_object_or_404(Turno, pk=turno_id)

    codigo_proveedor = turno.CodigoProveedor
    proveedores = {
        'AGODIF': 'DI FALCO MARIO DI FALCO JOSE Y DI FALCO COSME SOC DE HECHO',
        'BFBISE': 'BANCO MACRO S.A.',
        'OGPACK':'Tasky',
    }
    nombre_proveedor = proveedores.get(codigo_proveedor, 'Nombre Proveedor')

    # Asignar NombreProveedor AL OBJETO turno
    turno.NombreProveedor = nombre_proveedor

    timeline = [
        {
            'estado': 'Recepcionado',
            'completado': turno.Recepcionado,
            'fecha': turno.RecepcionadoFechaHora
        },
        {
            'estado': 'Auditado',
            'completado': turno.Auditado,
            'fecha': turno.AuditadoFechaHora
        },
        {
            'estado': 'Posicionado',
            'completado': turno.Posicionado,
            'fecha': turno.PosicionadoFechaHora
        }
    ]

    context = {
        'turno': turno,
        'timeline': timeline,
        'Nombre': 'Ver Turno'
    }

    return render(request, 'appConsultasTango/ver_turno.html', context)


def editar_turno(request, turno_id):
    turno = get_object_or_404(Turno, pk=turno_id)
    if request.method == 'POST':
        form = TurnoEditForm(request.POST, instance=turno)
        if form.is_valid():
            turno = form.save(commit=False)

            # Actualizar las fechas de los estados
            if turno.Recepcionado and not turno.RecepcionadoFechaHora:
                turno.RecepcionadoFechaHora = timezone.now()
            if turno.Auditado and not turno.AuditadoFechaHora:
                turno.AuditadoFechaHora = timezone.now()
            if turno.Posicionado and not turno.PosicionadoFechaHora:
                turno.PosicionadoFechaHora = timezone.now()

            turno.save()
            return redirect('herramientas:herramientas_ver_turno', turno_id=turno.IdTurno) # Updated redirect
    else:
        form = TurnoEditForm(instance=turno)

    return render(request, 'appConsultasTango/editar_turno.html', {'form': form, 'turno': turno, 'Nombre': 'Editar Turno'})

def lista_codigos_error(request):
    codigos = CodigosError.objects.all().order_by('CodigoError')
    nombre_template = 'Listado de Códigos de Error'
    return render(request, 'appConsultasTango/lista_codigos_error.html', {'codigos': codigos, 'Nombre': nombre_template})

def crear_codigo_error(request):
    if request.method == 'POST':
        form = CodigoErrorForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, 'Código de error creado exitosamente.')
            return redirect('herramientas:herramientas_lista_codigos_error') # Updated redirect
    else:
        form = CodigoErrorForm()
    return render(request, 'appConsultasTango/crear_editar_codigo_error.html', {'form': form, 'accion': 'Crear', 'Nombre': 'Crear Código de Error'})

def editar_codigo_error(request, codigo_id):
    codigo = get_object_or_404(CodigosError, CodigoError=codigo_id)
    if request.method == 'POST':
        form = CodigoErrorForm(request.POST, instance=codigo)
        if form.is_valid():
            form.save()
            messages.success(request, 'Código de error actualizado exitosamente.')
            return redirect('herramientas:herramientas_lista_codigos_error') # Updated redirect
    else:
        form = CodigoErrorForm(instance=codigo)
    return render(request, 'appConsultasTango/crear_editar_codigo_error.html', {'form': form, 'codigo': codigo, 'accion': 'Editar', 'Nombre': 'Editar Código de Error'})

def eliminar_codigo_error(request, codigo_id):
    try:
        codigo = CodigosError.objects.get(CodigoError=codigo_id)
    except CodigosError.DoesNotExist:
        messages.error(request, f'Código de error con ID {codigo_id} no encontrado.')
        return redirect('herramientas:herramientas_lista_codigos_error')
    except Exception as e:
        messages.error(request, f'Error al buscar el código: {e}')
        return redirect('herramientas:herramientas_lista_codigos_error')
    
    if request.method == 'POST':
        try:
            codigo.delete()
            messages.success(request, 'Código de error eliminado exitosamente.')
        except Exception as e:
            messages.error(request, f'Error al eliminar el código: {e}')
        return redirect('herramientas:herramientas_lista_codigos_error')
    
    context = {
        'codigo': codigo, 
        'Nombre': 'Eliminar Código de Error'
    }
    

# ============================================================================
# NUEVAS VISTAS PARA CALENDARIO DE RESERVAS DE TURNOS
# ============================================================================

@login_required(login_url="/login/")
def calendario_reservas(request):
    """
    Vista principal del calendario de reservas de turnos
    Muestra calendario interactivo con FullCalendar
    """
    # Obtener todos los estados activos para la leyenda
    estados = EstadoTurno.objects.filter(activo=True).order_by('orden_ejecucion')
    
    nombre_template = 'Calendario de Reservas'
    return render(request, 'appConsultasTango/calendario_reservas.html', {
        'Nombre': nombre_template,
        'estados': estados
    })


@login_required(login_url="/login/")
def obtener_turnos_calendario(request):
    """
    API endpoint para obtener turnos en formato FullCalendar
    """
    # Obtener parámetros de fecha del calendario
    start = request.GET.get('start')
    end = request.GET.get('end')
    
    try:
        # Convertir strings a objetos date
        if start:
            start_date = datetime.fromisoformat(start.replace('Z', '')).date()
        else:
            start_date = date.today()
            
        if end:
            end_date = datetime.fromisoformat(end.replace('Z', '')).date()
        else:
            end_date = start_date + timedelta(days=30)
        
        # Obtener turnos en el rango de fechas
        # Solo mostrar turnos con estados activos (no cancelados)
        turnos = TurnoReserva.objects.filter(
            fecha__gte=start_date,
            fecha__lte=end_date,
            estado__activo=True  # Solo estados activos
        ).select_related('estado')
        
        # Formatear para FullCalendar
        eventos = []
        for turno in turnos:
            # Combinar fecha y hora para crear datetime
            inicio = datetime.combine(turno.fecha, turno.hora_inicio)
            fin = datetime.combine(turno.fecha, turno.hora_fin)
            
            es_bloqueo_manual = turno.codigo_proveedor in ['HOT', 'INV', 'CYBER', 'ALTA']
            
            if es_bloqueo_manual:
                color = '#495057'  # Gris oscuro
                # Mostrar el nombre descriptivo directamente
                titulo = turno.nombre_proveedor or f"Bloqueo: {turno.codigo_proveedor}"
            else:
                color = turno.estado.color if turno.estado else '#17a2b8'
                titulo = f'{turno.codigo_proveedor} - OC: {formatear_orden_compra(turno.orden_compra)}'
            
            eventos.append({
                'id': turno.id_turno_reserva,
                'title': titulo,
                'start': inicio.isoformat(),
                'end': fin.isoformat(),
                'backgroundColor': color,
                'borderColor': color,
                'extendedProps': {
                    'codigo_proveedor': turno.codigo_proveedor,
                    'nombre_proveedor': turno.nombre_proveedor or '',
                    'orden_compra': formatear_orden_compra(turno.orden_compra),
                    'remitos': turno.remitos,
                    'cantidad_unidades': turno.cantidad_unidades,
                    'cantidad_bultos': turno.cantidad_bultos or 0,
                    'observaciones': turno.observaciones or '',
                    'estado': turno.estado.nombre if turno.estado else 'Sin Estado',
                    'estado_id': turno.estado.id_estado if turno.estado else None,
                    'estado_color': color,
                    'permite_editar': turno.estado.permite_editar if turno.estado else True,
                    'usuario_creador': turno.usuario_creador,
                    'isManualBlock': es_bloqueo_manual
                }
            })
        
        return JsonResponse(eventos, safe=False)
        
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=400)


@login_required(login_url="/login/")
def obtener_slots_disponibles(request):
    """
    API endpoint para obtener slots de tiempo disponibles en una fecha específica
    Retorna bloques de 30 minutos disponibles
    """
    fecha_str = request.GET.get('fecha')
    
    if not fecha_str:
        return JsonResponse({'error': 'Fecha requerida'}, status=400)
    
    try:
        fecha = datetime.strptime(fecha_str, '%Y-%m-%d').date()
        
        # Validar día de la semana
        weekday = fecha.weekday()
        if weekday == 6:  # Domingo
            return JsonResponse({'slots': []})
            
        if weekday == 5:  # Sábado
            hora_inicio = dt_time(7, 0)
            hora_fin = dt_time(10, 0)
        else:  # Lunes a Viernes
            hora_inicio = dt_time(7, 0)
            hora_fin = dt_time(16, 0)
        
        # Generar todos los slots de 30 minutos
        slots_disponibles = []
        current_time = datetime.combine(fecha, hora_inicio)
        end_time = datetime.combine(fecha, hora_fin)
        
        while current_time < end_time:
            slot_inicio = current_time.time()
            slot_fin = (current_time + timedelta(minutes=30)).time()
            
            # Filtrar breaks de comida para Lunes a Viernes
            if weekday < 5:  # L-V
                # Break 10:00 a 10:30
                if slot_inicio == dt_time(10, 0):
                    current_time += timedelta(minutes=30)
                    continue
                # Break 13:00 a 14:00 (cubre 13:00-13:30 y 13:30-14:00)
                if slot_inicio >= dt_time(13, 0) and slot_fin <= dt_time(14, 0):
                    current_time += timedelta(minutes=30)
                    continue
            
            # Verificar si el slot está ocupado
            # Solo considerar turnos con estados activos
            turnos_existentes = TurnoReserva.objects.filter(
                fecha=fecha,
                estado__activo=True
            )
            
            ocupado = False
            for turno in turnos_existentes:
                # Verificar superposición
                if not (slot_fin <= turno.hora_inicio or slot_inicio >= turno.hora_fin):
                    ocupado = True
                    break
            
            if not ocupado:
                slots_disponibles.append({
                    'inicio': slot_inicio.strftime('%H:%M'),
                    'fin': slot_fin.strftime('%H:%M'),
                    'display': f"{slot_inicio.strftime('%H:%M')} - {slot_fin.strftime('%H:%M')}"
                })
            
            current_time += timedelta(minutes=30)
        
        return JsonResponse({'slots': slots_disponibles})
        
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=400)


@login_required(login_url="/login/")
def nueva_reserva_turno(request):
    """
    Vista para crear una nueva reserva de turno
    Crea registro en historial de estados al crear
    Procesa archivos adjuntos si se envían
    """
    if request.method == 'POST':
        try:
            # Obtener el primer estado antes de crear el formulario
            primer_estado = EstadoTurno.objects.filter(activo=True).order_by('orden_ejecucion').first()
            
            # Crear una copia mutable de POST data
            post_data = request.POST.copy()
            
            # Si no viene estado en POST, agregar el primer estado
            if not post_data.get('estado') and primer_estado:
                post_data['estado'] = primer_estado.id_estado
            
            form = TurnoReservaForm(post_data, user=request.user)
            if form.is_valid():
                turno = form.save(commit=False)
                turno.usuario_creador = request.user.username
                
                # Asegurar que tiene estado (doble verificación)
                if not turno.estado and primer_estado:
                    turno.estado = primer_estado
                
                # Registrar datos de estado
                turno.usuario_ultima_modificacion_estado = request.user.username
                turno.estado_actual_desde = timezone.now()
                
                turno.save()
                
                # Si es un Turno Proveedor (Importado), actualizar la tabla RO_T_IMPORTACIONES_ENCABEZADO
                tipo_registro = post_data.get('tipo_registro')
                if tipo_registro == 'IMPORTADO':
                    try:
                        from datetime import datetime
                        dt_recibido = datetime.combine(turno.fecha, turno.hora_inicio)
                        
                        oc_list_str = turno.orden_compra
                        import ast
                        try:
                            oc_list = ast.literal_eval(oc_list_str)
                            if isinstance(oc_list, list) and oc_list:
                                oc_db_val = oc_list[0]
                            else:
                                oc_db_val = oc_list_str
                        except Exception:
                            oc_db_val = oc_list_str
                        
                        with connections['mi_db_2'].cursor() as cursor:
                            cursor.execute("""
                                UPDATE RO_T_IMPORTACIONES_ENCABEZADO 
                                SET FECHA_RECIBIDO = %s 
                                WHERE ORDEN_COMPRA = %s
                            """, [dt_recibido, oc_db_val])
                            connections['mi_db_2'].commit()
                    except Exception as e_up:
                        print(f"Error al actualizar FECHA_RECIBIDO: {str(e_up)}")
                
                # Crear registro en historial de estados
                HistorialEstadoTurno.objects.create(
                    turno=turno,
                    estado_anterior=None,  # Es el primer estado
                    estado_nuevo=turno.estado,
                    usuario=request.user.username,
                    observaciones="Creación de turno"
                )
                
                # Procesar archivos adjuntos si los hay
                archivos_subidos = 0
                archivos_error = []
                
                for key in request.FILES:
                    if key.startswith('archivo_'):
                        archivo = request.FILES[key]
                        tipo_key = 'tipo_documento_' + key.split('_')[1]
                        tipo_documento = request.POST.get(tipo_key, 'OTRO')
                        
                        try:
                            # Validar tamaño (5MB máximo)
                            if archivo.size > 5 * 1024 * 1024:
                                archivos_error.append(f'{archivo.name}: excede 5MB')
                                continue
                            
                            # Validar extensión
                            ext = os.path.splitext(archivo.name)[1].lower()
                            if ext not in EXTENSIONES_PERMITIDAS:
                                archivos_error.append(f'{archivo.name}: extensión no permitida')
                                continue
                            
                            # Detectar tipo MIME
                            tipo_mime, _ = mimetypes.guess_type(archivo.name)
                            if not tipo_mime:
                                tipo_mime = 'application/octet-stream'
                            
                            # Crear adjunto
                            AdjuntoTurnoReserva.objects.create(
                                turno=turno,
                                archivo=archivo,
                                tipo_documento=tipo_documento,
                                nombre_original=archivo.name,
                                tipo_archivo=tipo_mime,
                                tamaño_bytes=archivo.size,
                                usuario_subio=request.user.username
                            )
                            archivos_subidos += 1
                            
                        except Exception as e:
                            archivos_error.append(f'{archivo.name}: {str(e)}')
                
                mensaje = 'Turno reservado exitosamente'
                if archivos_subidos > 0:
                    mensaje += f'. Se adjuntaron {archivos_subidos} archivo(s).'
                if archivos_error:
                    mensaje += f' Errores en archivos: {", ".join(archivos_error)}'
                
                return JsonResponse({
                    'success': True, 
                    'message': mensaje,
                    'turno_id': turno.id_turno_reserva,
                    'archivos_subidos': archivos_subidos,
                    'archivos_error': archivos_error
                })
            else:
                # Retornar errores del formulario
                return JsonResponse({
                    'success': False,
                    'errors': form.errors
                }, status=400)
        except Exception as e:
            # Capturar cualquier excepción no prevista y retornar error descriptivo
            import traceback
            error_detail = str(e)
            traceback_detail = traceback.format_exc()
            
            # Log para debugging
            print(f"Error al crear turno: {error_detail}")
            print(traceback_detail)
            
            return JsonResponse({
                'success': False,
                'errors': {
                    '__all__': [f'Error del sistema: {error_detail}']
                }
            }, status=400)
    else:
        # GET request - mostrar formulario
        form = TurnoReservaForm(user=request.user)
        
        # Si viene fecha y hora por parámetro, pre-llenar
        fecha = request.GET.get('fecha')
        hora_inicio = request.GET.get('hora_inicio')
        hora_fin = request.GET.get('hora_fin')
        
        if fecha:
            form.fields['fecha'].initial = fecha
        if hora_inicio:
            form.fields['hora_inicio'].initial = hora_inicio
        if hora_fin:
            form.fields['hora_fin'].initial = hora_fin
    
    return render(request, 'appConsultasTango/nueva_reserva_turno.html', {
        'form': form,
        'Nombre': 'Nueva Reserva de Turno'
    })


@login_required(login_url="/login/")
def editar_reserva_turno(request, turno_id):
    """
    Vista para editar una reserva de turno existente
    Registra cambios de estado en historial
    Valida que no se puedan editar turnos de hoy o fechas pasadas
    Muestra en modo solo lectura si el estado no permite editar
    """
    turno = get_object_or_404(TurnoReserva, pk=turno_id)
    
    # Verificar si el estado actual permite editar el turno
    solo_lectura = turno.estado and not turno.estado.permite_editar
    
    # Verificar si la fecha del turno ya pasó
    hoy = date.today()
    turno_es_pasado = turno.fecha <= hoy
    
    if request.method == 'POST':
        # Rechazar POST si está en modo solo lectura
        if solo_lectura:
            messages.error(request, 'No se pueden realizar modificaciones en este turno.')
            return redirect('herramientas:herramientas_calendario_reservas')
        
        # Guardar estado y datos anteriores antes de aplicar cambios
        estado_anterior = turno.estado
        fecha_anterior = turno.fecha
        hora_inicio_anterior = turno.hora_inicio
        hora_fin_anterior = turno.hora_fin
        
        form = TurnoReservaForm(request.POST, instance=turno, user=request.user)
        if form.is_valid():
            turno_actualizado = form.save(commit=False)
            
            hubo_cambio_estado = estado_anterior != turno_actualizado.estado
            
            # Verificar si hubo cambio de fecha/horario
            cambios_detalles = []
            if fecha_anterior != turno_actualizado.fecha:
                cambios_detalles.append(f"Fecha: {fecha_anterior.strftime('%d/%m/%Y')} → {turno_actualizado.fecha.strftime('%d/%m/%Y')}")
            if hora_inicio_anterior != turno_actualizado.hora_inicio or hora_fin_anterior != turno_actualizado.hora_fin:
                cambios_detalles.append(f"Horario: {hora_inicio_anterior.strftime('%H:%M')}-{hora_fin_anterior.strftime('%H:%M')} → {turno_actualizado.hora_inicio.strftime('%H:%M')}-{turno_actualizado.hora_fin.strftime('%H:%M')}")
            
            if hubo_cambio_estado or cambios_detalles:
                observaciones_list = []
                if hubo_cambio_estado:
                    observaciones_list.append(f"Cambio de estado: {estado_anterior.nombre if estado_anterior else 'N/A'} → {turno_actualizado.estado.nombre}")
                    turno_actualizado.usuario_ultima_modificacion_estado = request.user.username
                    turno_actualizado.estado_actual_desde = timezone.now()
                if cambios_detalles:
                    observaciones_list.append(f"Modificación: {', '.join(cambios_detalles)}")
                
                # Crear registro en historial
                HistorialEstadoTurno.objects.create(
                    turno=turno_actualizado,
                    estado_anterior=estado_anterior,
                    estado_nuevo=turno_actualizado.estado,
                    usuario=request.user.username,
                    observaciones="; ".join(observaciones_list)
                )
            
            turno_actualizado.save()
            messages.success(request, 'Turno actualizado exitosamente.')
            # Redirigir a la misma página de edición para que el usuario vea el mensaje
            return redirect('herramientas:herramientas_editar_reserva_turno', turno_id=turno.id_turno_reserva)
        # Si hay errores, no agregar mensaje adicional ya que form.errors ya los muestra
    else:
        form = TurnoReservaForm(instance=turno, user=request.user)
    
    # Obtener historial de estados para mostrar en el template
    historial_estados = HistorialEstadoTurno.objects.filter(turno=turno).order_by('-fecha_cambio')
    
    # Obtener incidencias reportadas para este turno
    incidencias = IncidenciasTurno.objects.filter(turno=turno).select_related('codigo_error').order_by('-fecha_registro')
    
    # Obtener códigos de error activos para el modal
    codigos_error = CodigosError.objects.filter(Activo=True).order_by('Categoria', 'CodigoError')
    
    return render(request, 'appConsultasTango/editar_reserva_turno.html', {
        'form': form,
        'turno': turno,
        'turno_es_pasado': turno_es_pasado,
        'historial_estados': historial_estados,
        'incidencias': incidencias,
        'codigos_error': codigos_error,
        'solo_lectura': solo_lectura,
        'Nombre': f"{'Ver' if solo_lectura else 'Editar'} Turno #{turno.id_turno_reserva}"
    })


@login_required(login_url="/login/")
def eliminar_reserva_turno(request, turno_id):
    """
    Vista para eliminar/cancelar una reserva de turno
    """
    turno = get_object_or_404(TurnoReserva, pk=turno_id)
    
    if request.method == 'POST':
        try:
            # En lugar de eliminar, cambiar estado a CANCELADO
            turno.estado = 'CANCELADO'
            turno.save()
            messages.success(request, 'Turno cancelado exitosamente.')
            return JsonResponse({'success': True})
        except Exception as e:
            return JsonResponse({'success': False, 'error': str(e)}, status=400)
    
    return JsonResponse({'success': False, 'error': 'Método no permitido'}, status=405)


@login_required(login_url="/login/")
def detalle_reserva_turno(request, turno_id):
    """
    Vista para ver detalles de una reserva de turno
    """
    turno = get_object_or_404(TurnoReserva, pk=turno_id)
    
    return render(request, 'appConsultasTango/detalle_reserva_turno.html', {
        'turno': turno,
        'Nombre': 'Detalle de Reserva'
    })


@login_required(login_url="/login/")
@require_http_methods(["POST"])
def reportar_incidencia(request, turno_id):
    """
    Vista AJAX para reportar una incidencia en un turno
    """
    try:
        turno = get_object_or_404(TurnoReserva, pk=turno_id)
        codigo_error_id = request.POST.get('codigo_error')
        cantidad_afectada = request.POST.get('cantidad_afectada')
        detalle = request.POST.get('detalle')
        
        # Validar que se haya seleccionado un código de error
        if not codigo_error_id:
            return JsonResponse({'error': 'Debe seleccionar un código de error'}, status=400)
        
        # Obtener el código de error
        try:
            codigo_error = CodigosError.objects.get(CodigoError=codigo_error_id, Activo=True)
        except CodigosError.DoesNotExist:
            return JsonResponse({'error': 'El código de error seleccionado no es válido o está inactivo'}, status=400)
        
        # Crear la incidencia
        incidencia = IncidenciasTurno.objects.create(
            turno=turno,
            codigo_error=codigo_error,
            cantidad_afectada=int(cantidad_afectada) if cantidad_afectada else None,
            detalle=detalle if detalle else None,
            usuario_registro=request.user.username
        )
        
        return JsonResponse({
            'success': True,
            'message': 'Incidencia reportada exitosamente',
            'incidencia_id': incidencia.id_incidencia
        })
        
    except Exception as e:
        return JsonResponse({'error': f'Error al reportar incidencia: {str(e)}'}, status=500)


@login_required(login_url="/login/")
def listado_reservas(request):
    """
    Vista de listado de reservas (alternativa al calendario)
    """
    # Filtros opcionales
    fecha_desde = request.GET.get('fecha_desde')
    fecha_hasta = request.GET.get('fecha_hasta')
    estado = request.GET.get('estado')
    
    turnos = TurnoReserva.objects.all().select_related('estado')
    
    if fecha_desde:
        turnos = turnos.filter(fecha__gte=fecha_desde)
    if fecha_hasta:
        turnos = turnos.filter(fecha__lte=fecha_hasta)
    if estado:
        turnos = turnos.filter(estado=estado)
    
    turnos = turnos.order_by('-fecha', '-hora_inicio')
    
    estados = EstadoTurno.objects.filter(activo=True).order_by('orden_ejecucion')
    
    return render(request, 'appConsultasTango/listado_reservas.html', {
        'turnos': turnos,
        'estados': estados,
        'Nombre': 'Listado de Reservas'
    })


@login_required(login_url="/login/")
def historial_general_reservas(request):
    """
    Vista de historial general de cambios de reservas de turnos
    """
    historial = HistorialEstadoTurno.objects.all().select_related('turno', 'estado_anterior', 'estado_nuevo').order_by('-fecha_cambio')
    
    # Filtros opcionales
    turno_id = request.GET.get('turno_id')
    usuario = request.GET.get('usuario')
    if turno_id:
        historial = historial.filter(turno__id_turno_reserva=turno_id)
    if usuario:
        historial = historial.filter(usuario__icontains=usuario)
        
    # Paginación
    paginator = Paginator(historial, 50)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    return render(request, 'appConsultasTango/historial_general_reservas.html', {
        'Nombre': 'Historial de Cambios de Reservas',
        'page_obj': page_obj
    })


@login_required(login_url="/login/")
def descargar_reporte_reservas(request):
    """
    Genera y descarga un reporte Excel (.xlsx) con los turnos de reserva filtrados
    """
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    
    fecha_desde = request.GET.get('fecha_desde')
    fecha_hasta = request.GET.get('fecha_hasta')
    estado = request.GET.get('estado')
    
    turnos = TurnoReserva.objects.all().select_related('estado')
    
    if fecha_desde:
        turnos = turnos.filter(fecha__gte=fecha_desde)
    if fecha_hasta:
        turnos = turnos.filter(fecha__lte=fecha_hasta)
    if estado:
        turnos = turnos.filter(estado=estado)
        
    turnos = turnos.order_by('fecha', 'hora_inicio')
    
    # Crear Workbook de Excel
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Reporte de Reservas"
    
    # Estilos
    font_header = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
    fill_header = PatternFill(start_color='1F4E78', end_color='1F4E78', fill_type='solid') # Azul oscuro premium
    alignment_center = Alignment(horizontal='center', vertical='center', wrap_text=True)
    alignment_left = Alignment(horizontal='left', vertical='center')
    border_thin = Border(
        left=Side(style='thin', color='BFBFBF'),
        right=Side(style='thin', color='BFBFBF'),
        top=Side(style='thin', color='BFBFBF'),
        bottom=Side(style='thin', color='BFBFBF')
    )
    
    # Encabezados
    headers = [
        "ID", "Fecha", "Horario", "Código Prov.", "Proveedor / Tarea",
        "Orden de Compra", "Remitos", "Unidades", "Bultos", "Estado", "Creador"
    ]
    
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = header
        cell.font = font_header
        cell.fill = fill_header
        cell.alignment = alignment_center
        cell.border = border_thin
        
    # Datos
    for row_num, turno in enumerate(turnos, 2):
        # Formatear OC
        oc_str = formatear_orden_compra(turno.orden_compra)
        
        # Identificar si es bloqueo de tarea o turno regular
        nombre_prov = turno.nombre_proveedor or ''
        if turno.codigo_proveedor in ['HOT', 'INV', 'CYBER', 'ALTA']:
            nombre_prov = turno.nombre_proveedor or f"Bloqueo: {turno.codigo_proveedor}"
            
        data_row = [
            turno.id_turno_reserva,
            turno.fecha.strftime('%d/%m/%Y'),
            f"{turno.hora_inicio.strftime('%H:%M')} - {turno.hora_fin.strftime('%H:%M')}",
            turno.codigo_proveedor,
            nombre_prov,
            oc_str,
            turno.remitos,
            turno.cantidad_unidades,
            turno.cantidad_bultos or 0,
            turno.estado.nombre if turno.estado else 'Sin Estado',
            turno.usuario_creador
        ]
        
        for col_num, val in enumerate(data_row, 1):
            cell = ws.cell(row=row_num, column=col_num)
            cell.value = val
            cell.border = border_thin
            
            # Formatos de alineación y número
            if col_num in [1, 2, 3, 4, 8, 9, 10]:
                cell.alignment = alignment_center
            else:
                cell.alignment = alignment_left
                
            # Destacar bloqueos manuales
            if turno.codigo_proveedor in ['HOT', 'INV', 'CYBER', 'ALTA']:
                cell.fill = PatternFill(start_color='F2F2F2', end_color='F2F2F2', fill_type='solid')
                
    # Auto-ajustar ancho de columnas
    for col in ws.columns:
        max_len = 0
        col_letter = col[0].column_letter
        for cell in col:
            if cell.value:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = max(max_len + 3, 10)
        
    # Altura de filas
    ws.row_dimensions[1].height = 28
    for r in range(2, len(turnos) + 2):
        ws.row_dimensions[r].height = 20
        
    # Retornar como HttpResponse
    from django.http import HttpResponse
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    filename = f"reporte_reservas_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    
    wb.save(response)
    return response


@login_required(login_url="/login/")
def get_ordenes_compra_importadas(request):
    """
    Recupera las órdenes de compra de importación desde RO_T_IMPORTACIONES_ENCABEZADO
    que están pendientes de recepción (FECHA_RECIBIDO IS NULL).
    """
    q = request.GET.get('q', '').strip()
    if len(q) < 3:
        return JsonResponse({'ordenes': []})
        
    try:
        with connections['mi_db_2'].cursor() as cursor:
            # Filtrar por término de búsqueda (coincidencia en OC o Proveedor)
            cursor.execute("""
                SELECT DISTINCT ORDEN_COMPRA, PROVEEDOR, COD_PROVEE 
                FROM RO_T_IMPORTACIONES_ENCABEZADO 
                WHERE ORDEN_COMPRA IS NOT NULL 
                  AND FECHA_RECIBIDO IS NULL
                  AND (LTRIM(RTRIM(ORDEN_COMPRA)) LIKE %s OR PROVEEDOR LIKE %s)
                ORDER BY ORDEN_COMPRA
            """, [f'%{q}%', f'%{q}%'])
            rows = cursor.fetchall()
            
            ordenes = []
            for r in rows:
                oc_val = str(r[0]).strip()
                prov_val = str(r[1]).strip() if r[1] else 'Sin Proveedor'
                cod_provee = str(r[2]).strip() if r[2] else ''
                if oc_val:
                    ordenes.append({
                        'id': r[0],
                        'text': f"{oc_val} - {prov_val}",
                        'proveedor': prov_val,
                        'codigo_proveedor': cod_provee
                    })
            return JsonResponse({'ordenes': ordenes})
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)


# ============================================================================
# FIN VISTAS CALENDARIO DE RESERVAS
# ============================================================================

# ============================================================================
# FUNCIONES DE AYUDA PARA PERMISOS
# ============================================================================

def es_admin_o_logistica_sup(user):
    """Verificar si el usuario pertenece a Admin o Logistica_Sup"""
    if user.is_superuser:
        return True
    user_groups = [g.name for g in user.groups.all()]
    return 'Admin' in user_groups or 'Logistica_Sup' in user_groups

def es_admin_logistica_sup_o_logistica(user):
    """Verificar si el usuario pertenece a Admin, Logistica_Sup o Logistica"""
    if user.is_superuser:
        return True
    user_groups = [g.name for g in user.groups.all()]
    return 'Admin' in user_groups or 'Logistica_Sup' in user_groups or 'Logistica' in user_groups

# ============================================================================
# VISTAS CRUD PARA GESTIÓN DE ESTADOS DE TURNOS
# Solo accesibles para Admin y Logistica_Sup
# ============================================================================

@login_required(login_url="/login/")
@user_passes_test(es_admin_o_logistica_sup, login_url="/login/")
def listado_estados_turno(request):
    """
    Vista para listar todos los estados de turnos
    Solo accesible para Admin y Logistica_Sup
    """
    estados = EstadoTurno.objects.all().order_by('orden_ejecucion')
    
    return render(request, 'appConsultasTango/estados/listado_estados.html', {
        'estados': estados,
        'Nombre': 'Gestión de Estados de Turnos'
    })


@login_required(login_url="/login/")
@user_passes_test(es_admin_o_logistica_sup, login_url="/login/")
def crear_estado_turno(request):
    """
    Vista para crear un nuevo estado de turno
    Solo accesible para Admin y Logistica_Sup
    """
    if request.method == 'POST':
        form = EstadoTurnoForm(request.POST)
        if form.is_valid():
            estado = form.save()
            messages.success(request, f'Estado "{estado.nombre}" creado exitosamente.')
            return redirect('herramientas:herramientas_listado_estados_turno')
        else:
            messages.error(request, 'Error al crear el estado. Verifique los datos.')
    else:
        form = EstadoTurnoForm()
    
    return render(request, 'appConsultasTango/estados/crear_estado.html', {
        'form': form,
        'Nombre': 'Crear Estado de Turno'
    })


@login_required(login_url="/login/")
@user_passes_test(es_admin_o_logistica_sup, login_url="/login/")
def editar_estado_turno(request, estado_id):
    """
    Vista para editar un estado de turno existente
    Solo accesible para Admin y Logistica_Sup
    """
    estado = get_object_or_404(EstadoTurno, pk=estado_id)
    
    if request.method == 'POST':
        form = EstadoTurnoForm(request.POST, instance=estado)
        if form.is_valid():
            form.save()
            messages.success(request, f'Estado "{estado.nombre}" actualizado exitosamente.')
            return redirect('herramientas:herramientas_listado_estados_turno')
        else:
            messages.error(request, 'Error al actualizar el estado. Verifique los datos.')
    else:
        form = EstadoTurnoForm(instance=estado)
    
    return render(request, 'appConsultasTango/estados/editar_estado.html', {
        'form': form,
        'estado': estado,
        'Nombre': 'Editar Estado de Turno'
    })


@login_required(login_url="/login/")
@user_passes_test(es_admin_o_logistica_sup, login_url="/login/")
def eliminar_estado_turno(request, estado_id):
    """
    Vista para eliminar un estado de turno
    Solo accesible para Admin y Logistica_Sup
    Valida que no haya turnos usando este estado
    """
    estado = get_object_or_404(EstadoTurno, pk=estado_id)
    
    # Verificar si hay turnos usando este estado
    turnos_con_estado = TurnoReserva.objects.filter(estado=estado).count()
    
    if request.method == 'POST':
        if turnos_con_estado > 0:
            messages.error(
                request, 
                f'No se puede eliminar el estado "{estado.nombre}" porque hay {turnos_con_estado} turno(s) usándolo. '
                'Por favor, cambie el estado de esos turnos primero o desactive el estado en lugar de eliminarlo.'
            )
            return redirect('herramientas:herramientas_listado_estados_turno')
        
        try:
            nombre_estado = estado.nombre
            estado.delete()
            messages.success(request, f'Estado "{nombre_estado}" eliminado exitosamente.')
            return redirect('herramientas:herramientas_listado_estados_turno')
        except Exception as e:
            messages.error(request, f'Error al eliminar el estado: {str(e)}')
            return redirect('herramientas:herramientas_listado_estados_turno')
    
    return render(request, 'appConsultasTango/estados/eliminar_estado.html', {
        'estado': estado,
        'turnos_con_estado': turnos_con_estado,
        'Nombre': 'Eliminar Estado de Turno'
    })


@login_required(login_url="/login/")
@user_passes_test(es_admin_o_logistica_sup, login_url="/login/")
def reordenar_estados_turno(request):
    """
    Vista AJAX para reordenar estados mediante drag and drop
    Recibe orden_ejecucion para cada estado
    """
    if request.method == 'POST':
        try:
            import json
            data = json.loads(request.body)
            
            # data debe ser un array de objetos {id: X, orden: Y}
            for item in data:
                estado_id = item.get('id')
                nuevo_orden = item.get('orden')
                
                if estado_id and nuevo_orden:
                    EstadoTurno.objects.filter(pk=estado_id).update(orden_ejecucion=nuevo_orden)
            
            return JsonResponse({'success': True, 'message': 'Estados reordenados exitosamente'})
        except Exception as e:
            return JsonResponse({'success': False, 'error': str(e)}, status=400)
    
    return JsonResponse({'success': False, 'error': 'Método no permitido'}, status=405)


@login_required(login_url="/login/")
@user_passes_test(es_admin_o_logistica_sup, login_url="/login/")
def ejecutar_marcar_no_confirmados(request):
    """
    Vista para ejecutar manualmente el proceso de marcado de turnos NO CONFIRMADOS
    Solo Admin y Logistica_Sup pueden ejecutarla
    """
    if request.method == 'POST':
        try:
            from apps.home.SQL.Sql_Tango import marcar_turnos_no_confirmados
            
            turnos_marcados = marcar_turnos_no_confirmados()
            
            if turnos_marcados > 0:
                messages.success(
                    request,
                    f'Se marcaron {turnos_marcados} turno(s) como NO CONFIRMADO exitosamente.'
                )
            else:
                messages.info(request, 'No hay turnos RESERVADOS que deban marcarse como NO CONFIRMADO.')
            
        except Exception as e:
            messages.error(request, f'Error al ejecutar el proceso: {str(e)}')
        
        return redirect('herramientas:herramientas_listado_estados_turno')
    
    # GET request - mostrar confirmación
    return render(request, 'appConsultasTango/estados/confirmar_marcar_no_confirmados.html', {
        'Nombre': 'Marcar Turnos No Confirmados'
    })


# ============================================================================
# FIN VISTAS CRUD ESTADOS
# ============================================================================

    return render(request, 'appConsultasTango/confirmar_eliminar_codigo_error.html', context)

logger = logging.getLogger(__name__)

# @login_required(login_url="/login/")
# @method_decorator(csrf_exempt, name='dispatch')
class ImageUploadView(View):
    @method_decorator(login_required(login_url="/login/"))
    @method_decorator(csrf_exempt)
    def get(self, request):
        return render(request, 'appConsultasTango/uploadImg.html')

    @method_decorator(login_required(login_url="/login/"))
    @method_decorator(csrf_exempt)
    def post(self, request):
        if not request.FILES:
            logger.warning("No files received in POST request.")
            return JsonResponse({'error': 'No se recibió ningún archivo'}, status=400)

        file_key = next(iter(request.FILES), None)
        if not file_key:
            logger.warning("File key not found in request.FILES.")
            return JsonResponse({'error': 'Formato de archivo incorrecto o no se encontró el archivo'}, status=400)

        f = request.FILES.get(file_key)

        if not f:
            logger.warning(f"No file found for key: {file_key}")
            return JsonResponse({'error': 'No se encontró el archivo en la solicitud'}, status=400)

        try:
            file_path = os.path.join(settings.MEDIA_ROOT, 'imgTempEcommerce', f.name)
            counter = 1
            name, extension = os.path.splitext(f.name)
            while os.path.exists(file_path):
                file_path = os.path.join(settings.MEDIA_ROOT, 'imgTempEcommerce', f"{name}_{counter}{extension}")
                counter += 1

            with open(file_path, 'wb+') as destination:
                for chunk in f.chunks():
                    destination.write(chunk)

            logger.info(f"Successfully saved file: {f.name} to {file_path}")

            # --- CAMBIO AQUÍ: Asegurar que siempre se incluya la información del archivo ---
            # return JsonResponse({
            #     'message': 'Imagen cargada con éxito',
            #     'filename': f.name, # Envía el nombre del archivo específico
            #     'filepath': file_path.replace(settings.MEDIA_ROOT, '') # Opcional: ruta relativa si es útil
            # })
            # --- CAMBIO AQUÍ: Simplificar la respuesta JSON para Dropzone ---
            # Dropzone es bastante flexible con respuestas 200, un JSON vacío
            # o simple a menudo funciona mejor para marcar el éxito.
            return JsonResponse({'message': 'success'}, status=200) # O simplemente {}
            # return JsonResponse({}, status=200) # Esto también debería funcionar

        except Exception as e:
            logger.error(f"Error processing file {f.name}: {e}", exc_info=True)
            return JsonResponse({'error': f'Error al procesar la imagen {f.name}', 'details': str(e)}, status=500)

@login_required(login_url="/login/")
def upload_success(request):
    return render(request, 'appConsultasTango/successImg.html')

# ***************************
# ***************************

@login_required(login_url="/login/")
def Eliminar_Turno(request):
    if request.method == 'POST':
        IdTurno = request.POST.get('IdTurno')
        datos = Turno.objects.get(IdTurno=IdTurno)
        datos.delete()
        # Redirigir a la página de éxito
        return redirect('herramientas:herramientas_listar_turno') # Updated redirect
    else:
        IdTurno = request.GET.get('IdTurno')
        datos = Turno.objects.get(IdTurno=IdTurno)
        return render(request, 'appConsultasTango/Eliminar_turno.html', {'datos': datos})

@login_required(login_url="/login/")
def Editar_Turno(request,IdTurno):
    datos = Turno.objects.get(IdTurno=IdTurno)
    if request.method == 'POST':
        form = TurnoForm(request.POST or None, request.FILES or None, instance=datos)
        if form.is_valid():
            form.save()
            # Redirigir a la página de éxito
            return redirect('herramientas:herramientas_listar_turno') # Updated redirect
    else:
        form = TurnoForm(instance=datos)
    return render(request, 'appConsultasTango/Editar_turno.html', {'form': form})

@login_required(login_url="/login/")
def Listar_turno(request):
    datos = Turno.objects.all()
    return render(request,'appConsultasTango/turno_list.html', {'turnos': datos})

@login_required(login_url="/login/")
def Crear_turno(request):
    if request.method == 'POST':
        form = TurnoForm(request.POST)
        if form.is_valid():
            form.save()
            # Redirigir a la página de éxito
            return redirect('herramientas:herramientas_listar_turno') # Updated redirect
    else:
        form = TurnoForm()
    return render(request, 'appConsultasTango/Crear_turno.html', {'form': form})


@login_required(login_url="/login/")
def Gestion_cronograma(request):
    Nombre = 'Gestión cronograma'
    dir_iframe = DIR_HERAMIENTAS['Gestion_cronograma']
    return render(request, 'home/PlantillaHerramientas.html', {'dir_iframe': dir_iframe, })


@login_required(login_url="/login/")
def Gestion_guias_mayoristas(request):
    Nombre = 'Guías mayoristas'
    dir_iframe = DIR_HERAMIENTAS['Gestion_guias_mayoristas']
    return redirect(dir_iframe)

@login_required(login_url="/login/")
def ImpRotulos(request):
    Nombre = 'Imprecion de Rotulos'
    dir_iframe = DIR_HERAMIENTAS['ImpRotulos']
    # return redirect(dir_iframe)
    return render(request, 'home/PlantillaHerramientas.html', {'dir_iframe': dir_iframe, })

@login_required(login_url="/login/")
def ImpRemEcom(request):
    Nombre = ''
    dir_iframe = DIR_HERAMIENTAS['ImportarRemEcommerce']
    # return redirect(dir_iframe)
    return render(request, 'home/PlantillaHerramientas.html', {'dir_iframe': dir_iframe, })

@login_required(login_url="/login/")
def SistemaReclamos(request):
    Nombre = 'Sistema de Reclamos'
    dir_iframe = DIR_HERAMIENTAS['SistemaReclamos']
    return redirect(dir_iframe)



# Abastecimiento

@login_required(login_url="/login/")
def StockBase(request):
    Nombre = 'Stock Base'
    dir_iframe = DIR_HERAMIENTAS['StockBase']
    return render(request, 'home/PlantillaHerramientas.html', {'dir_iframe': dir_iframe,'Nombre':Nombre})

@login_required(login_url="/login/")
def MedidasLocales(request):
    Nombre = 'Medidas Locales'
    dir_iframe = DIR_HERAMIENTAS['MedidasLocales']
    return render(request, 'home/PlantillaHerramientas.html', {'dir_iframe': dir_iframe,'Nombre':Nombre})

@login_required(login_url="/login/")
def Recodificacion(request):
    Nombre = 'Recodificacion Outlet'
    dir_iframe = DIR_HERAMIENTAS['Recodificacion']
    return render(request, 'home/PlantillaHerramientas.html', {'dir_iframe': dir_iframe,'Nombre':Nombre})

@login_required(login_url="/login/")
def Stock_excluido(request):
    Nombre = 'Stock excluido'
    dir_iframe = DIR_HERAMIENTAS['Stock_excluido']
    return render(request, 'home/PlantillaHerramientas.html', {'dir_iframe': dir_iframe, })


@login_required(login_url="/login/")
def MaestroDestinos(request):
    Nombre = 'MaestroDestinos'
    dir_iframe = DIR_HERAMIENTAS['MaestroDestinos']
    return redirect(dir_iframe)

@login_required(login_url="/login/")
def GestionEquivalentes(request):
    Nombre = ''
    dir_iframe = DIR_HERAMIENTAS['GestionEquivalentes']
    return redirect(dir_iframe)

# Comercial
@login_required(login_url="/login/")
def Gestion_categoria_productos(request):
    Nombre = ''
    dir_iframe = DIR_HERAMIENTAS['Gestion_categoria_productos'] #+ UserName
    return render(request, 'home/PlantillaHerramientas.html', {'dir_iframe': dir_iframe,'Nombre':Nombre })

@login_required(login_url="/login/")
def AdministrarCuotas(request):
    Nombre = 'Administrar Cuotas'
    dir_iframe = DIR_HERAMIENTAS['AdministrarCuotas'] #+ UserName
    return render(request, 'home/PlantillaHerramientas.html', {'dir_iframe': dir_iframe,'Nombre':Nombre })

@login_required(login_url="/login/")
def AdministrarInternos(request):
    Nombre = 'Administrar Internos'
    dir_iframe = DIR_HERAMIENTAS['AdministrarInternos'] #+ UserName
    return render(request, 'home/PlantillaHerramientas.html', {'dir_iframe': dir_iframe,'Nombre':Nombre })


@login_required(login_url="/login/")
def PromoBancos(request):
    Nombre = 'Promo Bancos'
    dir_iframe = DIR_HERAMIENTAS['PromoBancos'] #+ UserName
    return render(request, 'home/PlantillaHerramientas.html', {'dir_iframe': dir_iframe,'Nombre':Nombre })

@login_required(login_url="/login/")
def AltaNuevosLocales(request):
    Nombre = 'Alta de Nuevos Locales'
    dir_iframe = DIR_HERAMIENTAS['AltaNuevosLocales'] #+ UserName
    # return render(request, 'home/PlantillaHerramientas.html', {'dir_iframe': dir_iframe,'Nombre':Nombre })
    return redirect(dir_iframe)

@login_required(login_url="/login/")
def UsuariosFranquicias(request):
    Nombre = 'Usuarios de Franquicias'
    dir_iframe = DIR_HERAMIENTAS['UsuariosFranquicias'] #+ UserName
    return render(request, 'home/PlantillaHerramientas.html', {'dir_iframe': dir_iframe,'Nombre':Nombre })

@login_required(login_url="/login/")
def ObjetivosVentaFranquicias(request):
    Nombre = 'Objetivos Ventas Franquicias'
    dir_iframe = DIR_HERAMIENTAS['ObjetivosVentaFranquicias'] #+ UserName
    return render(request, 'home/PlantillaHerramientas.html', {'dir_iframe': dir_iframe,'Nombre':Nombre })

@login_required(login_url="/login/")
def gestionKits(request):
    Nombre = 'Gestión de Kits'
    dir_iframe = DIR_HERAMIENTAS['gestionKits'] #+ UserName
    return render(request, 'home/PlantillaHerramientas.html', {'dir_iframe': dir_iframe,'Nombre':Nombre })

@login_required(login_url="/login/")
def conversorCSV(request):
    if request.method == 'POST':
        uploaded = request.FILES.get('excel_file')
        if not uploaded:
            return render(request, 'herramientas/conversor_csv.html', {'error': 'No se recibió ningún archivo.'})

        separator = request.POST.get('separator', ';')
        encoding  = request.POST.get('encoding', 'utf-8-sig')
        sheet     = request.POST.get('sheet_name', 0)  # 0 = primera hoja

        try:
            df = pd.read_excel(uploaded, sheet_name=sheet, dtype=str)
        except Exception as e:
            return render(request, 'herramientas/conversor_csv.html', {'error': f'Error al leer el archivo: {e}'})

        import io
        buffer = io.StringIO()
        df.to_csv(buffer, sep=separator, index=False, encoding=encoding)
        csv_content = buffer.getvalue()

        # Nombre del archivo de salida
        base = uploaded.name.rsplit('.', 1)[0]
        response = HttpResponse(csv_content.encode(encoding), content_type='text/csv')
        response['Content-Disposition'] = f'attachment; filename="{base}.csv"'
        return response

    return render(request, 'herramientas/conversor_csv.html')


@login_required(login_url="/login/")
def conversorCSV_sheets(request):
    """Devuelve la lista de hojas de un archivo Excel (llamada AJAX)."""
    if request.method == 'POST':
        uploaded = request.FILES.get('excel_file')
        if not uploaded:
            return JsonResponse({'sheets': []})
        try:
            xl = pd.ExcelFile(uploaded)
            return JsonResponse({'sheets': xl.sheet_names})
        except Exception:
            return JsonResponse({'sheets': []})
    return JsonResponse({'sheets': []})

# Mayoristas
@login_required(login_url="/login/")
def gestionPedidos(request):
    Nombre = 'Gestión de Pedidos Mayoristas'
    dir_iframe = DIR_HERAMIENTAS['gestionPedidos']
    return redirect(dir_iframe)

@login_required(login_url="/login/")
def ActualizacionPrecios(request):
    Nombre='Actualización de Precios'
    dir_iframe = DIR_HERAMIENTAS['ActualizacionPrecios']
    return render(request,'home/PlantillaHerramientas.html',{'dir_iframe':dir_iframe,'Nombre':Nombre})
# Ecommerce

# --- Categorías ---
@login_required(login_url="/login/")
def categoria_list(request):
    categorias = obtener_categorias()
    paginator = Paginator(categorias, 10)  # Muestra 10 registros por página
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    return render(request, 'appConsultasTango/ecommerce/categorias/categoria_list.html', {'page_obj': page_obj})

@login_required(login_url="/login/")
def categoria_create(request):
    if request.method == 'POST':
        form = CategoriaForm(request.POST)
        if form.is_valid():
            crear_categoria(form.cleaned_data['nombre'], form.cleaned_data['codigo'], form.cleaned_data['PalabrasClave'])
            messages.success(request, 'Categoría creada exitosamente.')
            return redirect('herramientas:herramientas_categoria_list') # Updated redirect
    else:
        form = CategoriaForm()
    return render(request, 'appConsultasTango/ecommerce/categorias/categoria_create.html', {'form': form})

@login_required(login_url="/login/")
def categoria_update(request, id_categoria):
    categoria = obtener_categoria(id_categoria)
    if not categoria:
        messages.error(request, 'Categoría no encontrada.')
        return redirect('herramientas:herramientas_categoria_list') # Updated redirect
    if request.method == 'POST':
        form = CategoriaForm(request.POST)
        if form.is_valid():
            editar_categoria(id_categoria, form.cleaned_data['nombre'], form.cleaned_data['codigo'], form.cleaned_data['PalabrasClave'])
            messages.success(request, 'Categoría actualizada exitosamente.')
            return redirect('herramientas:herramientas_categoria_list') # Updated redirect
    else:
        form = CategoriaForm(initial={'nombre': categoria[1], 'codigo': categoria[2], 'PalabrasClave': categoria[3]})
    return render(request, 'appConsultasTango/ecommerce/categorias/categoria_update.html', {'form': form, 'id_categoria': id_categoria})

@login_required(login_url="/login/")
def categoria_delete(request, id_categoria):
    categoria = obtener_categoria(id_categoria)
    if not categoria:
        messages.error(request, 'Categoría no encontrada.')
        return redirect('herramientas:herramientas_categoria_list') # Updated redirect
    if request.method == 'POST':
         eliminar_categoria(id_categoria)
         messages.success(request, 'Categoría eliminada exitosamente.')
         return redirect('herramientas:herramientas_categoria_list') # Updated redirect
    return render(request, 'appConsultasTango/ecommerce/categorias/categoria_delete.html', {'categoria': categoria})

# --- Subcategorías ---
@login_required(login_url="/login/")
def subcategoria_list(request):
    subcategorias = obtener_subcategorias()
    # paginator = Paginator(subcategorias, 10)
    # page_number = request.GET.get('page')
    # page_obj = paginator.get_page(page_number)
    return render(request, 'appConsultasTango/ecommerce/subcategorias/subcategoria_list.html', {'page_obj': subcategorias})


@login_required(login_url="/login/")
def subcategoria_create(request):
    if request.method == 'POST':
        form = SubcategoriaForm(request.POST)
        if form.is_valid():
            crear_subcategoria(form.cleaned_data['codigo'], form.cleaned_data['nombre'],
                               form.cleaned_data['Keywords'], form.cleaned_data['id_categoria_VtxAr'],form.cleaned_data['id_categoria_Tango'])
            messages.success(request, 'Subcategoría creada exitosamente.')
            return redirect('herramientas:herramientas_subcategoria_list') # Updated redirect
    else:
        form = SubcategoriaForm()
    return render(request, 'appConsultasTango/ecommerce/subcategorias/subcategoria_create.html', {'form': form})

@login_required(login_url="/login/")
def subcategoria_update(request, id_subcategoria):
    subcategoria = obtener_subcategoria(id_subcategoria)
    if not subcategoria:
        messages.error(request, 'Subcategoría no encontrada.')
        return redirect('herramientas:herramientas_subcategoria_list') # Updated redirect
    if request.method == 'POST':
        form = SubcategoriaForm(request.POST)
        if form.is_valid():
             editar_subcategoria(id_subcategoria, form.cleaned_data['codigo'], form.cleaned_data['nombre'],
                               form.cleaned_data['Keywords'], form.cleaned_data['id_categoria_VtxAr'],form.cleaned_data['id_categoria_Tango'])

             messages.success(request, 'Subcategoría actualizada exitosamente.')
             return redirect('herramientas:herramientas_subcategoria_list') # Updated redirect
    else:
          form = SubcategoriaForm(initial={'codigo': subcategoria[1],'nombre': subcategoria[2],'Keywords': subcategoria[3],'id_categoria_VtxAr': subcategoria[4],'id_categoria_Tango': subcategoria[5]})
    return render(request, 'appConsultasTango/ecommerce/subcategorias/subcategoria_update.html', {'form': form, 'id_subcategoria': id_subcategoria})

@login_required(login_url="/login/")
def subcategoria_delete(request, id_subcategoria):
    subcategoria = obtener_subcategoria(id_subcategoria)
    if not subcategoria:
        messages.error(request, 'Subcategoría no encontrada.')
        return redirect('herramientas:herramientas_subcategoria_list') # Updated redirect
    if request.method == 'POST':
         eliminar_subcategoria(id_subcategoria)
         messages.success(request, 'Subcategoría eliminada exitosamente.')
         return redirect('herramientas:herramientas_subcategoria_list') # Updated redirect
    return render(request, 'appConsultasTango/ecommerce/subcategorias/subcategoria_delete.html', {'subcategoria': subcategoria})

# --- Relaciones ---
@login_required(login_url="/login/")
def relacion_list(request):
    relaciones = obtener_relaciones()
    # paginator = Paginator(relaciones, 10)
    # page_number = request.GET.get('page')
    # page_obj = paginator.get_page(page_number)
    return render(request, 'appConsultasTango/ecommerce/relaciones/relacion_list.html', {'page_obj': relaciones})

@login_required(login_url="/login/")
def relacion_create(request):
    if request.method == 'POST':
        form = RelacionForm(request.POST)
        if form.is_valid():
            crear_relacion(form.cleaned_data['id_categoria_Tango'], form.cleaned_data['id_subCat_VtxAr'])
            messages.success(request, 'Relación creada exitosamente.')
            return redirect('herramientas:herramientas_relacion_list') # Updated redirect
    else:
        form = RelacionForm()
    return render(request, 'appConsultasTango/ecommerce/relaciones/relacion_create.html', {'form': form})


@login_required(login_url="/login/")
def relacion_update(request, id_categoria_tango,id_subcategoria):
    relacion = obtener_relacion(id_categoria_tango,id_subcategoria)
    if not relacion:
        messages.error(request, 'Relación no encontrada.')
        return redirect('herramientas:herramientas_relacion_create') # Updated redirect
    if request.method == 'POST':
        form = RelacionForm(request.POST)
        if form.is_valid():
             editar_relacion(id_categoria_tango,id_subcategoria,form.cleaned_data['id_categoria_Tango'], form.cleaned_data['id_subCat_VtxAr'])
             messages.success(request, 'Relación actualizada exitosamente.')
             return redirect('herramientas:herramientas_relacion_list') # Updated redirect
    else:
        form = RelacionForm(initial={'id_categoria_Tango': relacion[0], 'id_subCat_VtxAr': relacion[1]})
    return render(request, 'appConsultasTango/ecommerce/relaciones/relacion_update.html', {'form': form, 'id_categoria_tango': id_categoria_tango,'id_subcategoria':id_subcategoria})


@login_required(login_url="/login/")
def relacion_delete(request, id_categoria_tango, id_subcategoria):
    relacion = obtener_relacion(id_categoria_tango,id_subcategoria)
    if not relacion:
        messages.error(request, 'Relación no encontrada.')
        return redirect('herramientas:herramientas_relacion_list') # Updated redirect
    if request.method == 'POST':
        eliminar_relacion(id_categoria_tango, id_subcategoria)
        messages.success(request, 'Relación eliminada exitosamente.')
        return redirect('herramientas:herramientas_relacion_list') # Updated redirect
    return render(request, 'appConsultasTango/ecommerce/relaciones/relacion_delete.html', {'relacion': relacion,'id_categoria_tango': id_categoria_tango, 'id_subcategoria': id_subcategoria})

@login_required(login_url="/login/")
def import_art_vtex(request):
    nombre_db='LAKER_SA'
    settings.DATABASES['mi_db_2']['NAME'] = nombre_db
    print('Cambiando base de datos a ' + nombre_db)
    print('import_art_vtex')
    try:
        if request.method == 'POST' and request.FILES['excel_file']:
            pfile = request.FILES['excel_file']
            filesys =FileSystemStorage()
            # Obtener el nombre del archivo sin espacios en blanco
            filename = pfile.name.replace(' ', '')

            # Guardar el archivo con el nombre sin espacios en blanco
            uploadfilename = filesys.save(filename, pfile)
            extension = os.path.splitext(uploadfilename)
            if  not extension[1] == '.xlsx':
                error_extension = 'El formato del archivo debe ser de tipo .xlsx'
                os.remove(filesys.path(uploadfilename))
                return render(request,'appConsultasTango/importFileArtVtex.html',{'mensaje_error':error_extension})

            uploaded_url = filesys.url(uploadfilename)  #Ruta donde se guardo el archivo
            uploaded_url = os.path.normpath(uploaded_url)
            # print("uploaded_url: " + uploaded_url)
            ruta_actual = os.path.join(os.getcwd()) #Ruta del proyecto
            # print('Ruta actual: ' + ruta_actual)
            path_filname = ruta_actual + uploaded_url
            wb = openpyxl.load_workbook(path_filname) #Abrimos el archivo
            worksheet = wb.active
            excel_data = list()
            enc_data = list()
            mensaje_error = ''
            mensaje_Success = ''
            art_valido = ''
            nombre_archivo = "AltaArtVtex.xls"
            eliminar_archivo_excel(nombre_archivo)
            # iterando sobre las filas y obteniendo
            # valor de cada celda en la fila
            for row in worksheet.iter_rows():
                fila = row[0].row   #Numero de fila
                row_data = list()
                talon_pedido = 0
                i = 1
                if worksheet.cell(row=fila, column=1).value is None: #Frena el loop al llegar al final de la lista
                    break

                for cell in row:
                    if fila == 1:
                        enc_data.append(str(cell.value))
                    else:
                        if worksheet.cell(row=1, column=i).value == 'ARTICULO':
                            if fila > 1:
                                articulo = worksheet.cell(row=fila, column=i).value
                                art_valido= validar_articulo(articulo)
                                print('art_valido: ' + str(art_valido))
                                if art_valido == 'ERROR':
                                    mensaje_error = 'Hay articulos que no existen en SJ_ETIQUETAS_FINAL'
                                    row_data.append('*' + str(cell.value) + '*')
                                    i += 1
                                    continue
                        elif worksheet.cell(row=1, column=i).value == 'DESCRIPCION':
                            worksheet.cell(row=fila, column=i).value = art_valido

                        if cell.value == None:
                            row_data.append(str(''))
                        else:
                            row_data.append(str(cell.value))
                    i += 1

                excel_data.append(row_data)
                if len(row_data) > 0:
                    # Validar que row_data tenga suficientes elementos
                    if len(row_data) >= 3:
                        resultado_json = obtenerInformacionArticulo(row_data[0], row_data[2])
                        if resultado_json is not None:
                            try:
                                resultado = json.loads(resultado_json)
                                # Validar que el resultado no sea None y sea una lista válida
                                if resultado is not None and isinstance(resultado, list) and len(resultado) > 0:
                                    crear_archivo_excel(resultado, nombre_archivo)
                                else:
                                    print(f"El stored procedure devolvió datos inválidos para artículo {row_data[0]}: {resultado_json}")
                                    if not mensaje_error:
                                        mensaje_error = f'El stored procedure devolvió datos inválidos para el artículo {row_data[0]}'
                            except json.JSONDecodeError as e:
                                print(f"Error decodificando JSON para artículo {row_data[0]}: {str(e)}")
                                print(f"Contenido recibido: {resultado_json}")
                                if not mensaje_error:
                                    mensaje_error = 'Error procesando algunos artículos. Revisar datos de entrada.'
                        else:
                            print(f"No se pudo obtener información para el artículo: {row_data[0]}")
                            if not mensaje_error:
                                mensaje_error = 'Error obteniendo información de algunos artículos.'
                    else:
                        print(f"Fila con datos insuficientes: {row_data}")
                        if not mensaje_error:
                            mensaje_error = 'Algunas filas del archivo tienen datos insuficientes.'

            wb.save(path_filname)
            if not(mensaje_error):
                mensaje_Success = 'Articulos cargados correctamente'
                os.remove(filesys.path(uploadfilename))
            else:
                os.remove(filesys.path(uploadfilename))

            return render(request, 'appConsultasTango/importFileArtVtex.html' ,{'enc_data':enc_data,'excel_data':excel_data,'mensaje_Success':mensaje_Success,'mensaje_error':mensaje_error})


    except Exception as identifier:
        print(f"Error en import_art_vtex: {str(identifier)}")
        mensaje_error = f"Error procesando el archivo: {str(identifier)}"
        return render(request, 'appConsultasTango/importFileArtVtex.html', {
            'mensaje_error': mensaje_error
        })
    return render(request,'appConsultasTango/importFileArtVtex.html',{})

import xlwt
import xlrd

def crear_archivo_excel(tempJson, nombre_archivo):
    # Validar que tempJson no sea None y sea una lista válida
    if tempJson is None:
        print(f"Error: tempJson es None - no se puede crear el archivo {nombre_archivo}")
        return
    
    if not isinstance(tempJson, list) or len(tempJson) == 0:
        print(f"Error: tempJson no es una lista válida o está vacía - no se puede crear el archivo {nombre_archivo}")
        return

    # Construir la ruta absoluta del archivo
    ruta_archivo = os.path.join(settings.MEDIA_ROOT, nombre_archivo)

    try:
        # Abrir el archivo de Excel existente
        libro_rd = xlrd.open_workbook(ruta_archivo)
        hoja_rd = libro_rd.sheet_by_index(0)
        ultima_fila = hoja_rd.nrows

        # Crear un nuevo objeto Workbook
        libro_wt = xlwt.Workbook()
        hoja_wt = libro_wt.add_sheet('Hoja 1')

        # Copiar los datos de la hoja existente a la nueva hoja
        for i in range(ultima_fila):
            for j in range(hoja_rd.ncols):
                hoja_wt.write(i, j, hoja_rd.cell_value(i, j))

        # Agregar los nuevos datos a la hoja
        for i, fila in enumerate(tempJson):
            for j, valor in enumerate(fila.values()):
                hoja_wt.write(ultima_fila + i, j, valor)

        # Guardar el libro en un archivo
        libro_wt.save(ruta_archivo)
    except FileNotFoundError:
        # Si el archivo no existe, crear uno nuevo
        libro_wt = xlwt.Workbook()
        hoja_wt = libro_wt.add_sheet('Hoja 1')

        # Agregar los encabezados si es el primer registro
        if len(tempJson) > 0:
            encabezados = list(tempJson[0].keys())
            for i, encabezado in enumerate(encabezados):
                hoja_wt.write(0, i, encabezado)

        # Agregar los datos a la hoja
        for i, fila in enumerate(tempJson):
            for j, valor in enumerate(fila.values()):
                hoja_wt.write(i + 1, j, valor)

        # Guardar el libro en un archivo
        libro_wt.save(ruta_archivo)

def eliminar_archivo_excel(nombre_archivo):
    # Construir la ruta absoluta del archivo
    ruta_archivo = os.path.join(settings.MEDIA_ROOT, nombre_archivo)
    print('ruta_archivo: ' + ruta_archivo)
    # Verificar si el archivo existe y eliminarlo si es así
    if os.path.exists(ruta_archivo):
        os.remove(ruta_archivo)
        print('Se elimino el archivo ' + nombre_archivo)

@login_required(login_url="/login/")
def Control_pedidos(request):
    Nombre = 'Control pedidos'
    dir_iframe = DIR_HERAMIENTAS['Control_pedidos']
    return render(request, 'home/PlantillaHerramientas.html', {'dir_iframe': dir_iframe, })


@login_required(login_url="/login/")
def StockSegVtex(request):
    Nombre = 'Adm. Stock Seguridad Vtex'
    dir_iframe = DIR_HERAMIENTAS['StockSegVtex']
    return render(request, 'home/PlantillaHerramientas.html', {'dir_iframe': dir_iframe, })

@login_required(login_url="/login/")
def novICBC(request):
    Nombre = 'Actualizar novedades ICBC'
    dir_iframe = DIR_HERAMIENTAS['novICBC']
    # return render(request, 'home/PlantillaHerramientas.html', {'dir_iframe': dir_iframe,'Nombre':Nombre})
    return redirect(dir_iframe)

# Gerencia
@login_required(login_url="/login/")
def rendircobranzas(request,UserName):
    Nombre = 'Rendir Cobranzas'
    dir_iframe = DIR_HERAMIENTAS['rendircobranzas'] + UserName
    return render(request, 'home/PlantillaHerramientas.html', {'dir_iframe': dir_iframe,'Nombre':Nombre })

@login_required(login_url="/login/")
def GestionarCobro(request,UserName):
    Nombre = 'Gestionar Cobro'
    dir_iframe = DIR_HERAMIENTAS['gestionarCobro'] + UserName
    return render(request, 'home/PlantillaHerramientas.html', {'dir_iframe': dir_iframe,'Nombre':Nombre })

@login_required(login_url="/login/")
def RegistrarEfectivo(request,UserName):
    Nombre = 'Registrar Efectivo'
    dir_iframe = DIR_HERAMIENTAS['registrarEfectivo'] + UserName
    return render(request, 'home/PlantillaHerramientas.html', {'dir_iframe': dir_iframe,'Nombre':Nombre })

@login_required(login_url="/login/")
def gestionPremiosComercial(request):
    Nombre = ''
    dir_iframe = DIR_HERAMIENTAS['gestionPremiosComercial']
    return render(request, 'home/PlantillaHerramientas.html', {'dir_iframe': dir_iframe,'Nombre':Nombre })

# Administracion

@login_required(login_url="/login/")
def EgresosCajaTesoreria(request):
    Nombre = ''
    dir_iframe = DIR_HERAMIENTAS['EgresosCajaTesoreria'] #+ UserName
    return render(request, 'home/PlantillaHerramientas.html', {'dir_iframe': dir_iframe,'Nombre':Nombre })

@login_required(login_url="/login/")
def CargaFacturasSuc(request):
    Nombre = 'Carga Facturas Suc'
    dir_iframe = DIR_HERAMIENTAS['CargaFacturasSuc'] #+ UserName
    return render(request, 'home/PlantillaHerramientas.html', {'dir_iframe': dir_iframe,'Nombre':Nombre })


@login_required(login_url="/login/")
def ControlGastosSupervision(request):
    Nombre = 'Gastos Supervision'
    dir_iframe = DIR_HERAMIENTAS['ControlGastosSupervision'] #+ UserName
    return render(request, 'home/PlantillaHerramientas.html', {'dir_iframe': dir_iframe,'Nombre':Nombre })

@login_required(login_url="/login/")
def ControlMasivoCobranza(request):
    Nombre = 'Control Gastos'
    dir_iframe = DIR_HERAMIENTAS['ControlMasivoCobranza'] #+ UserName
    return redirect(dir_iframe)

@login_required(login_url="/login/")
def Controlgastos(request):
    Nombre = 'Control Gastos'
    dir_iframe = DIR_HERAMIENTAS['controlGastos'] #+ UserName
    return redirect(dir_iframe)

@login_required(login_url="/login/")
def Cargargastos(request):
    Nombre = 'Gestionar Cobro'
    dir_iframe = DIR_HERAMIENTAS['cargaGastos'] #+ UserName
    return redirect(dir_iframe)

@login_required(login_url="/login/")
def Controlcajasdiario(request):
    Nombre = 'Control cajas Diario'
    dir_iframe = DIR_HERAMIENTAS['Controlcajasdiario'] #+ UserName
    return render(request, 'home/PlantillaHerramientas.html', {'dir_iframe': dir_iframe,'Nombre':Nombre })

@login_required(login_url="/login/")
def CargaGastosAlquileres(request):
    Nombre = 'Carga Gastos Alquileres'
    dir_iframe = DIR_HERAMIENTAS['CargaGastosAlquileres'] #+ UserName
    return redirect(dir_iframe)

@login_required(login_url="/login/")
def ControlEgresosDeCaja(request,UserName):
    Nombre = 'Control Egresos De Caja'
    dir_iframe = DIR_HERAMIENTAS['ControlEgresosDeCaja'] + UserName
    return redirect(dir_iframe)

@login_required(login_url="/login/")
def CargarContratosDeAlquiler(request):
    Nombre = 'Cargar Contratos De Alquiler'
    dir_iframe = DIR_HERAMIENTAS['CargarContratosDeAlquiler'] #+ UserName
    return redirect(dir_iframe)

@login_required(login_url="/login/")
def RelacionesCtaCont(request):
    Nombre = ''
    dir_iframe = DIR_HERAMIENTAS['RelacionesCtaCont'] #+ UserName
    return redirect(dir_iframe)

@login_required(login_url="/login/")
def GestionDeProveedores(request):
    Nombre = ''
    dir_iframe = DIR_HERAMIENTAS['GestionDeProveedores'] #+ UserName
    return redirect(dir_iframe)

@login_required(login_url="/login/")
def RegistroPagoServicios(request):
    Nombre = ''
    dir_iframe = DIR_HERAMIENTAS['RegistroPagoServicios'] #+ UserName
    return render(request, 'home/PlantillaHerramientas.html', {'dir_iframe': dir_iframe,'Nombre':Nombre })

# Administracion_CE             ***Comercio Exterior***
@login_required(login_url="/login/")
def Cargarcontenedor(request):
    Nombre = 'Gestión de Importaciones'
    dir_iframe = DIR_HERAMIENTAS['cargaInicial'] #+ UserName
    return redirect(dir_iframe)

@login_required(login_url="/login/")
def EditarContenedor(request):
    Nombre = 'Editar Contenedor'
    dir_iframe = DIR_HERAMIENTAS['mostrarOrden'] #+ UserName
    return render(request, 'home/PlantillaHerramientas.html', {'dir_iframe': dir_iframe,'Nombre':Nombre })

@login_required(login_url="/login/")
def VentasLocatarios(request):
    Nombre = 'Ventas Locatarios'
    dir_iframe = DIR_HERAMIENTAS['VentasLocatarios'] #+ UserName
    return redirect(dir_iframe)

@login_required(login_url="/login/")
def FacturasDirectores(request):
    Nombre = 'Facturas Directores'
    dir_iframe = DIR_HERAMIENTAS['FacturasDirectores'] #+ UserName
    return render(request, 'home/PlantillaHerramientas.html', {'dir_iframe': dir_iframe,'Nombre':Nombre })

# RRHH

@login_required(login_url="/login/")
def CargaAnticipo(request):
    Nombre = 'Carga de anticipos'
    dir_iframe = DIR_HERAMIENTAS['CargaAnticipo']
    return render(request, 'home/PlantillaHerramientas.html', {'dir_iframe': dir_iframe,'Nombre':Nombre})

@login_required(login_url="/login/")
def altaVendedores(request):
    Nombre = ''
    dir_iframe = DIR_HERAMIENTAS['altaVendedores'] #+ UserName
    # return render(request, 'home/PlantillaHerramientas.html', {'dir_iframe': dir_iframe,'Nombre':Nombre })
    return redirect(dir_iframe)

@login_required(login_url="/login/")
def listarGrupos(request):
    Nombre = ''
    dir_iframe = DIR_HERAMIENTAS['listarGrupos'] #+ UserName
    return render(request, 'home/PlantillaHerramientas.html', {'dir_iframe': dir_iframe,'Nombre':Nombre })

@login_required(login_url="/login/")
def gestionarVendedores(request):
    Nombre = ''
    dir_iframe = DIR_HERAMIENTAS['gestionarVendedores'] #+ UserName
    # return render(request, 'home/PlantillaHerramientas.html', {'dir_iframe': dir_iframe,'Nombre':Nombre })
    return redirect(dir_iframe)

# @login_required(login_url="/login/")
# def adminEmpleados(request):
#     Nombre = ''
#     dir_iframe = DIR_HERAMIENTAS['adminEmpleados'] #+ UserName
#     return render(request, 'home/PlantillaHerramientas.html', {'dir_iframe': dir_iframe,'Nombre':Nombre })
#     # return redirect(dir_iframe)

# Tesoreria

@login_required(login_url="/login/")
def ControlDeEfectivo(request):
    Nombre = ''
    dir_iframe = DIR_HERAMIENTAS['ControlDeEfectivo'] #+ UserName
    return render(request, 'home/PlantillaHerramientas.html', {'dir_iframe': dir_iframe,'Nombre':Nombre })
    # return redirect(dir_iframe)

@login_required(login_url="/login/")
def PagosDirectores(request):
    Nombre = 'Pagos a Directores'
    dir_iframe = DIR_HERAMIENTAS['PagosDirectores'] #+ UserName
    return render(request, 'home/PlantillaHerramientas.html', {'dir_iframe': dir_iframe,'Nombre':Nombre })
    # return redirect(dir_iframe)    

# Supervisores

@login_required(login_url="/login/")
def CargaProyecto(request):
    Nombre = ''
    dir_iframe = DIR_HERAMIENTAS['CargaProyecto'] #+ UserName
    return render(request, 'home/PlantillaHerramientas.html', {'dir_iframe': dir_iframe,'Nombre':Nombre })
    # return redirect(dir_iframe)

# ─────────────────────────────────────────────────────────────────────────────
# Gestión de Usuarios
# ─────────────────────────────────────────────────────────────────────────────
from django.contrib.auth.models import User, Group

def _puede_gestionar_usuarios(user):
    """Retorna True si el usuario puede acceder a la gestión de usuarios."""
    if not user.is_authenticated:
        return False
    return user.groups.filter(name='admin').exists() or \
           user.groups.filter(name__iendswith='_sup').exists()


def _grupos_accesibles(user):
    """Retorna el queryset de grupos que el usuario puede asignar.
    Soporta múltiples grupos _sup (ej: Abastecimiento_Sup + Comercial_sup)."""
    if user.groups.filter(name='admin').exists():
        return Group.objects.all().order_by('name')
    from django.db.models import Q
    sup_groups = user.groups.filter(name__iendswith='_sup')
    if not sup_groups.exists():
        return Group.objects.none()
    q = Q()
    for sg in sup_groups:
        prefijo = sg.name[:-4]  # quita '_sup' o '_Sup'
        q |= Q(name__istartswith=prefijo)
    return Group.objects.filter(q).exclude(name__iendswith='_sup').order_by('name')


def _usuarios_accesibles(user, solo_activos=True):
    """Retorna el queryset de usuarios que el usuario puede ver/editar.
    Siempre incluye usuarios activos sin grupos asignados para que cualquier
    operador pueda asignarles un grupo al momento de darlos de alta.
    """
    from django.db.models import Q
    qs = User.objects.prefetch_related('groups')
    if solo_activos:
        qs = qs.filter(is_active=True)
    if user.groups.filter(name='admin').exists():
        return qs.order_by('username')
    grupos = _grupos_accesibles(user)
    # Usuarios del área del operador + usuarios activos sin ningún grupo asignado
    return qs.filter(
        Q(groups__in=grupos) | Q(groups__isnull=True)
    ).distinct().order_by('username')


@login_required(login_url="/login/")
def gestion_usuarios(request):
    if not _puede_gestionar_usuarios(request.user):
        from django.http import HttpResponseForbidden
        return HttpResponseForbidden("No tiene permiso para acceder a esta sección.")
    is_admin = request.user.groups.filter(name='admin').exists()
    grupos_disponibles = list(_grupos_accesibles(request.user).values('id', 'name'))
    return render(request, 'herramientas/gestion_usuarios.html', {
        'is_admin': is_admin,
        'grupos_disponibles': grupos_disponibles,
    })


@login_required(login_url="/login/")
def api_usuarios(request):
    """Retorna lista de usuarios en JSON."""
    if not _puede_gestionar_usuarios(request.user):
        return JsonResponse({'error': 'Sin permiso'}, status=403)

    solo_activos = request.GET.get('activos', 'true') == 'true'
    qs = _usuarios_accesibles(request.user, solo_activos=solo_activos)

    data = []
    for u in qs:
        grupos = [g.name for g in u.groups.all()]
        data.append({
            'id': u.id,
            'username': u.username,
            'first_name': u.first_name,
            'last_name': u.last_name,
            'email': u.email,
            'is_active': u.is_active,
            'date_joined': u.date_joined.strftime('%d/%m/%Y'),
            'last_login': u.last_login.strftime('%d/%m/%Y %H:%M') if u.last_login else '—',
            'grupos': grupos,
        })
    return JsonResponse({'usuarios': data})


@login_required(login_url="/login/")
def api_usuario_detalle(request, user_id):
    """Retorna detalle de un usuario para edición."""
    if not _puede_gestionar_usuarios(request.user):
        return JsonResponse({'error': 'Sin permiso'}, status=403)

    target = get_object_or_404(User, pk=user_id)

    # Verificar que el solicitante tiene acceso a este usuario
    accesibles = _usuarios_accesibles(request.user, solo_activos=False)
    if not accesibles.filter(pk=user_id).exists():
        return JsonResponse({'error': 'Sin permiso sobre este usuario'}, status=403)

    grupos_usuario = list(target.groups.values_list('id', flat=True))
    # Grupos que el operador puede gestionar normalmente
    grupos_accesibles_ids = set(_grupos_accesibles(request.user).values_list('id', flat=True))
    # Grupos que el usuario tiene pero están fuera del alcance del operador
    grupos_extra = list(
        target.groups.exclude(id__in=grupos_accesibles_ids).values('id', 'name').order_by('name')
    )
    return JsonResponse({
        'id': target.id,
        'username': target.username,
        'first_name': target.first_name,
        'last_name': target.last_name,
        'email': target.email,
        'is_active': target.is_active,
        'grupos': grupos_usuario,
        'grupos_extra': grupos_extra,  # [{id, name}] asignados fuera del área del operador
    })


@login_required(login_url="/login/")
def api_editar_usuario(request, user_id):
    """Edita datos personales y grupos de un usuario."""
    if request.method != 'POST':
        return JsonResponse({'error': 'Método no permitido'}, status=405)
    if not _puede_gestionar_usuarios(request.user):
        return JsonResponse({'error': 'Sin permiso'}, status=403)

    target = get_object_or_404(User, pk=user_id)
    accesibles = _usuarios_accesibles(request.user, solo_activos=False)
    if not accesibles.filter(pk=user_id).exists():
        return JsonResponse({'error': 'Sin permiso sobre este usuario'}, status=403)

    try:
        data = json.loads(request.body)
    except json.JSONDecodeError:
        return JsonResponse({'error': 'Datos inválidos'}, status=400)

    first_name = data.get('first_name', target.first_name).strip()
    last_name = data.get('last_name', target.last_name).strip()
    email = data.get('email', target.email).strip()
    is_active = data.get('is_active', target.is_active)
    if not isinstance(is_active, bool):
        is_active = bool(is_active)
    grupos_ids = data.get('grupos', [])

    # Grupos que el operador puede gestionar por su área
    grupos_permitidos = _grupos_accesibles(request.user)
    grupos_permitidos_ids = set(grupos_permitidos.values_list('id', flat=True))
    # Grupos extra: los que el usuario ya tiene asignados fuera del área del operador.
    # El operador los ve en la UI y puede desasignarlos.
    grupos_usuario_actuales_ids = set(target.groups.values_list('id', flat=True))
    grupos_extra_ids = grupos_usuario_actuales_ids - grupos_permitidos_ids
    # El operador puede manejar: su área normal + los grupos extra ya asignados
    grupos_manejables_ids = grupos_permitidos_ids | grupos_extra_ids
    grupos_ids_validos = [int(gid) for gid in grupos_ids if int(gid) in grupos_manejables_ids]

    target.first_name = first_name
    target.last_name = last_name
    target.email = email
    target.is_active = is_active
    target.save(update_fields=['first_name', 'last_name', 'email', 'is_active'])

    # Preservar solo grupos que el operador no pudo ver en absoluto (ninguno en este caso,
    # pero se mantiene como red de seguridad por si hubiera grupos no incluidos en la UI).
    grupos_a_preservar = target.groups.exclude(id__in=grupos_manejables_ids)
    nuevos_grupos = list(grupos_a_preservar.values_list('id', flat=True)) + grupos_ids_validos
    target.groups.set(nuevos_grupos)

    return JsonResponse({'ok': True, 'mensaje': f'Usuario {target.username} actualizado correctamente.'})


@login_required(login_url="/login/")
def api_cambiar_password(request, user_id):
    """Cambia la contraseña de un usuario."""
    if request.method != 'POST':
        return JsonResponse({'error': 'Método no permitido'}, status=405)
    if not _puede_gestionar_usuarios(request.user):
        return JsonResponse({'error': 'Sin permiso'}, status=403)

    target = get_object_or_404(User, pk=user_id)
    accesibles = _usuarios_accesibles(request.user, solo_activos=False)
    if not accesibles.filter(pk=user_id).exists():
        return JsonResponse({'error': 'Sin permiso sobre este usuario'}, status=403)

    try:
        data = json.loads(request.body)
    except json.JSONDecodeError:
        return JsonResponse({'error': 'Datos inválidos'}, status=400)

    nueva_password = data.get('password', '').strip()
    confirmar_password = data.get('confirmar_password', '').strip()

    if not nueva_password:
        return JsonResponse({'error': 'La contraseña no puede estar vacía.'}, status=400)
    if len(nueva_password) < 6:
        return JsonResponse({'error': 'La contraseña debe tener al menos 6 caracteres.'}, status=400)
    if nueva_password != confirmar_password:
        return JsonResponse({'error': 'Las contraseñas no coinciden.'}, status=400)

    target.set_password(nueva_password)
    target.save(update_fields=['password'])

    return JsonResponse({'ok': True, 'mensaje': f'Contraseña de {target.username} actualizada correctamente.'})


# Admin



# --- Moved from viewsExtras.py ---
@login_required(login_url="/login/")
def import_file_etiquetas(request):
    nombre_db='LAKER_SA'
    settings.DATABASES['mi_db_2']['NAME'] = nombre_db
    print('Cambiando base de datos a ' + nombre_db)
    print('import_file_etiquetas')
    try:
        if request.method == 'POST' and request.FILES['excel_file']:
            pfile = request.FILES['excel_file']
            filesys =FileSystemStorage()
            # Obtener el nombre del archivo sin espacios en blanco
            filename = pfile.name.replace(' ', '')

            # Guardar el archivo con el nombre sin espacios en blanco
            uploadfilename = filesys.save(filename, pfile)
            extension = os.path.splitext(uploadfilename)
            if  not extension[1] == '.xlsx':
                error_extension = 'El formato del archivo debe ser de tipo .xlsx'
                os.remove(filesys.path(uploadfilename))
                return render(request,'appConsultasTango/importFileEtiquetas.html',{'mensaje_error':error_extension})

            uploaded_url = filesys.url(uploadfilename)  #Ruta donde se guardo el archivo
            uploaded_url = os.path.normpath(uploaded_url)
            # print("uploaded_url: " + uploaded_url)
            ruta_actual = os.path.join(os.getcwd()) #Ruta del proyecto
            # print('Ruta actual: ' + ruta_actual)
            path_filname = ruta_actual + uploaded_url
            wb = openpyxl.load_workbook(path_filname) #Abrimos el archivo
            worksheet = wb.active
            excel_data = list()
            enc_data = list()
            mensaje_error = ''
            mensaje_Success = ''
            art_valido = ''
            # iterando sobre las filas y obteniendo
            # valor de cada celda en la fila
            for row in worksheet.iter_rows():
                fila = row[0].row   #Numero de fila
                row_data = list()
                talon_pedido = 0
                i = 1
                if worksheet.cell(row=fila, column=1).value is None: #Frena el loop al llegar al final de la lista
                    break

                for cell in row:
                    if fila == 1:
                        enc_data.append(str(cell.value))
                    else:
                        if worksheet.cell(row=1, column=i).value == 'ARTICULO':
                            if fila > 1:
                                articulo = worksheet.cell(row=fila, column=i).value
                                art_valido= validar_articulo(articulo)
                                if art_valido == 'ERROR':
                                    mensaje_error = 'Hay articulos que no existen en SJ_ETIQUETAS_FINAL'
                                    row_data.append('*' + str(cell.value) + '*')
                                    i += 1
                                    continue
                        elif worksheet.cell(row=1, column=i).value == 'DESCRIPCION':
                            worksheet.cell(row=fila, column=i).value = art_valido

                        if cell.value == None:
                            row_data.append(str(''))
                        else:
                            row_data.append(str(cell.value))
                    i += 1

                excel_data.append(row_data)

            # print('path_filname: ' + path_filname)
            wb.save(path_filname)
            if not(mensaje_error):
                upload_file_artEtiquetas(path_filname) #Carga articulos en base de datos
                mensaje_Success = 'Articulos cargados correctamente'
                os.remove(filesys.path(uploadfilename))

            else:
                os.remove(filesys.path(uploadfilename))


            return render(request, 'appConsultasTango/importFileEtiquetas.html' ,{'enc_data':enc_data,'excel_data':excel_data,'mensaje_Success':mensaje_Success,'mensaje_error':mensaje_error})


    except Exception as identifier:
        print(identifier)
    return render(request,'appConsultasTango/importFileEtiquetas.html',{})

def upload_file_artEtiquetas(path_filname):
    excel_file =path_filname
    empexceldata = pd.read_excel(excel_file, engine='openpyxl')
    dbframe = empexceldata
    borrar_contTabla('SJ_T_ETIQUETAS_FINAL')

    for df in dbframe.itertuples():
        articulo = str(df.ARTICULO)
        descripcion = str(df.DESCRIPCION)
        cargar_articulo(articulo,descripcion)

@login_required(login_url="/login/")
def import_file_anularPedidos(request):
    nombre_db='LAKER_SA'
    settings.DATABASES['mi_db_2']['NAME'] = nombre_db
    print('Cambiando base de datos a ' + nombre_db)
    try:
        if request.method == 'POST' and request.FILES.get('excel_file'):
            pfile = request.FILES['excel_file']
            filesys =FileSystemStorage()
            # Obtener el nombre del archivo sin espacios en blanco
            filename = pfile.name.replace(' ', '')

            # Guardar el archivo con el nombre sin espacios en blanco
            uploadfilename = filesys.save(filename, pfile)
            extension = os.path.splitext(uploadfilename)
            if  not extension[1] == '.xlsx':
                error_extension = '❌ El archivo debe ser formato .xlsx (Excel). Formato recibido: ' + extension[1]
                os.remove(filesys.path(uploadfilename))
                return render(request,'appConsultasTango/importFilePedidosAnular.html',{'mensaje_error':error_extension})

            # Obtener la ruta completa del archivo
            path_filname = filesys.path(uploadfilename)
            print(f'Ruta del archivo: {path_filname}')
            
            wb = openpyxl.load_workbook(path_filname) #Abrimos el archivo
            worksheet = wb.active
            excel_data = list()
            enc_data = list()
            pedidos_data = list()  # Lista estructurada de pedidos
            mensaje_error = ''
            mensaje_info = ''
            pedidos_validos = 0
            pedidos_invalidos = 0
            
            # Buscar índices de columnas
            col_nro_pedido = None
            col_talon_ped = None
            for i, cell in enumerate(worksheet[1], start=1):
                if cell.value == 'NRO_PEDIDO':
                    col_nro_pedido = i
                elif cell.value == 'TALON_PED':
                    col_talon_ped = i
            
            if not col_nro_pedido or not col_talon_ped:
                os.remove(filesys.path(uploadfilename))
                return render(request,'appConsultasTango/importFilePedidosAnular.html',{
                    'mensaje_error': '❌ El archivo debe contener las columnas NRO_PEDIDO y TALON_PED'
                })
            
            # Procesar filas
            for row in worksheet.iter_rows(min_row=2):
                if not row[0].value:  # Fila vacía
                    break
                
                nro_pedido = row[col_nro_pedido - 1].value
                talon_ped = row[col_talon_ped - 1].value
                
                if nro_pedido is None or talon_ped is None:
                    continue
                
                # Validar pedido
                ped_valido = validar_pedido(nro_pedido, str(talon_ped))
                estado = 'válido' if ped_valido > 0 else 'inválido'
                
                pedido_info = {
                    'nro_pedido': str(nro_pedido),
                    'talon_ped': str(talon_ped),
                    'estado': estado,
                    'valido': ped_valido > 0
                }
                pedidos_data.append(pedido_info)
                
                if ped_valido > 0:
                    pedidos_validos += 1
                else:
                    pedidos_invalidos += 1
            
            os.remove(filesys.path(uploadfilename))
            
            # Preparar mensaje informativo
            if pedidos_validos > 0 and pedidos_invalidos == 0:
                mensaje_info = f'📋 Se encontraron {pedidos_validos} pedido(s) válido(s) listo(s) para anular'
            elif pedidos_validos > 0 and pedidos_invalidos > 0:
                mensaje_info = f'⚠️ Se encontraron {pedidos_validos} pedido(s) válido(s) y {pedidos_invalidos} inválido(s). Solo se pueden anular los válidos.'
            elif pedidos_validos == 0:
                mensaje_error = f'❌ No se encontraron pedidos válidos para anular. Total de pedidos inválidos: {pedidos_invalidos}'
            
            # Si es petición AJAX, devolver JSON
            if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                return JsonResponse({
                    'success': True,
                    'pedidos_data': pedidos_data,
                    'pedidos_validos': pedidos_validos,
                    'pedidos_invalidos': pedidos_invalidos,
                    'mensaje_info': mensaje_info,
                    'mensaje_error': mensaje_error
                })
            
            # SOLO PREVISUALIZACIÓN - NO SE ANULA NADA AÚN
            return render(request, 'appConsultasTango/importFilePedidosAnular.html', {
                'pedidos_data': pedidos_data,
                'pedidos_validos': pedidos_validos,
                'pedidos_invalidos': pedidos_invalidos,
                'mensaje_info': mensaje_info,
                'mensaje_error': mensaje_error,
                'mostrar_preview': True
            })

    except Exception as identifier:            
        print(identifier)
        mensaje_error = f'❌ Error al procesar el archivo: {str(identifier)}. Verifica que el archivo tenga el formato correcto.'
        # Si es petición AJAX, devolver JSON de error
        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            return JsonResponse({'success': False, 'mensaje_error': mensaje_error})
        return render(request,'appConsultasTango/importFilePedidosAnular.html',{'mensaje_error': mensaje_error})
    return render(request,'appConsultasTango/importFilePedidosAnular.html',{})

@login_required(login_url="/login/")
def import_file_anularPedidosUY(request):
    nombre_db='TASKY_SA'
    settings.DATABASES['mi_db_2']['NAME'] = nombre_db
    print('Cambiando base de datos a ' + nombre_db)
    try:
        if request.method == 'POST' and request.FILES.get('excel_file'):
            pfile = request.FILES['excel_file']
            filesys =FileSystemStorage()
            # Obtener el nombre del archivo sin espacios en blanco
            filename = pfile.name.replace(' ', '')

            # Guardar el archivo con el nombre sin espacios en blanco
            uploadfilename = filesys.save(filename, pfile)
            extension = os.path.splitext(uploadfilename)
            if  not extension[1] == '.xlsx':
                error_extension = '❌ El archivo debe ser formato .xlsx (Excel). Formato recibido: ' + extension[1]
                os.remove(filesys.path(uploadfilename))
                return render(request,'appConsultasTango/importFilePedidosAnularUY.html',{'mensaje_error':error_extension})

            # Obtener la ruta completa del archivo
            path_filname = filesys.path(uploadfilename)
            print(f'Ruta del archivo (UY): {path_filname}')
            
            wb = openpyxl.load_workbook(path_filname) #Abrimos el archivo
            worksheet = wb.active
            pedidos_data = list()  # Lista estructurada de pedidos
            mensaje_error = ''
            mensaje_info = ''
            pedidos_validos = 0
            pedidos_invalidos = 0
            
            # Buscar índices de columnas
            col_nro_pedido = None
            col_talon_ped = None
            for i, cell in enumerate(worksheet[1], start=1):
                if cell.value == 'NRO_PEDIDO':
                    col_nro_pedido = i
                elif cell.value == 'TALON_PED':
                    col_talon_ped = i
            
            if not col_nro_pedido or not col_talon_ped:
                os.remove(filesys.path(uploadfilename))
                return render(request,'appConsultasTango/importFilePedidosAnularUY.html',{
                    'mensaje_error': '❌ El archivo debe contener las columnas NRO_PEDIDO y TALON_PED'
                })
            
            # Procesar filas
            for row in worksheet.iter_rows(min_row=2):
                if not row[0].value:  # Fila vacía
                    break
                
                nro_pedido = row[col_nro_pedido - 1].value
                talon_ped = row[col_talon_ped - 1].value
                
                if nro_pedido is None or talon_ped is None:
                    continue
                
                # Validar pedido
                ped_valido = validar_pedido(nro_pedido, str(talon_ped))
                estado = 'válido' if ped_valido > 0 else 'inválido'
                
                pedido_info = {
                    'nro_pedido': str(nro_pedido),
                    'talon_ped': str(talon_ped),
                    'estado': estado,
                    'valido': ped_valido > 0
                }
                pedidos_data.append(pedido_info)
                
                if ped_valido > 0:
                    pedidos_validos += 1
                else:
                    pedidos_invalidos += 1
            
            os.remove(filesys.path(uploadfilename))
            
            # Preparar mensaje informativo
            if pedidos_validos > 0 and pedidos_invalidos == 0:
                mensaje_info = f'📋 Se encontraron {pedidos_validos} pedido(s) válido(s) listo(s) para anular (Uruguay)'
            elif pedidos_validos > 0 and pedidos_invalidos > 0:
                mensaje_info = f'⚠️ Se encontraron {pedidos_validos} pedido(s) válido(s) y {pedidos_invalidos} inválido(s). Solo se pueden anular los válidos.'
            elif pedidos_validos == 0:
                mensaje_error = f'❌ No se encontraron pedidos válidos para anular. Total de pedidos inválidos: {pedidos_invalidos}'
            
            # Si es petición AJAX, devolver JSON
            if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                return JsonResponse({
                    'success': True,
                    'pedidos_data': pedidos_data,
                    'pedidos_validos': pedidos_validos,
                    'pedidos_invalidos': pedidos_invalidos,
                    'mensaje_info': mensaje_info,
                    'mensaje_error': mensaje_error
                })
            
            # SOLO PREVISUALIZACIÓN - NO SE ANULA NADA AÚN
            return render(request, 'appConsultasTango/importFilePedidosAnularUY.html', {
                'pedidos_data': pedidos_data,
                'pedidos_validos': pedidos_validos,
                'pedidos_invalidos': pedidos_invalidos,
                'mensaje_info': mensaje_info,
                'mensaje_error': mensaje_error,
                'mostrar_preview': True,
                'es_uy': True
            })

    except Exception as identifier:            
        print(identifier)
        mensaje_error = f'❌ Error al procesar el archivo: {str(identifier)}. Verifica que el archivo tenga el formato correcto.'
        # Si es petición AJAX, devolver JSON de error
        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            return JsonResponse({'success': False, 'mensaje_error': mensaje_error})
        return render(request,'appConsultasTango/importFilePedidosAnularUY.html',{'mensaje_error': mensaje_error})
    return render(request,'appConsultasTango/importFilePedidosAnularUY.html',{})

@login_required(login_url="/login/")
def ejecutar_anulacion_pedidos_uy(request):
    """Vista para ejecutar la anulación de pedidos seleccionados (Uruguay)"""
    nombre_db='TASKY_SA'
    settings.DATABASES['mi_db_2']['NAME'] = nombre_db
    
    if request.method == 'POST':
        try:
            # Obtener pedidos seleccionados del POST
            pedidos_seleccionados = request.POST.getlist('pedidos_seleccionados')
            anular_todos = request.POST.get('anular_todos') == 'true'
            
            if not pedidos_seleccionados:
                return render(request, 'appConsultasTango/importFilePedidosAnularUY.html', {
                    'mensaje_error': '❌ No se seleccionaron pedidos para anular'
                })
            
            # Procesar anulaciones
            anulados_exitosos = 0
            anulados_fallidos = 0
            
            for pedido_str in pedidos_seleccionados:
                try:
                    # Formato: "nro_pedido|talon_ped"
                    nro_pedido, talon_ped = pedido_str.split('|')
                    numero_pedido = ' 0000' + str(nro_pedido)[-9:]
                    
                    # Ejecutar anulación
                    anular_pedido(talon_ped, numero_pedido)
                    anulados_exitosos += 1
                except Exception as e:
                    print(f"Error anulando pedido {pedido_str}: {str(e)}")
                    anulados_fallidos += 1
            
            # Preparar mensaje de resultado
            if anulados_fallidos == 0:
                mensaje_success = f'✅ Se anularon exitosamente {anulados_exitosos} pedido(s) (Uruguay)'
            else:
                mensaje_success = f'⚠️ Se anularon {anulados_exitosos} pedido(s). Fallaron {anulados_fallidos}.'
            
            return render(request, 'appConsultasTango/importFilePedidosAnularUY.html', {
                'mensaje_Success': mensaje_success
            })
            
        except Exception as e:
            print(f"Error en ejecutar_anulacion_pedidos_uy: {str(e)}")
            return render(request, 'appConsultasTango/importFilePedidosAnularUY.html', {
                'mensaje_error': f'❌ Error al anular pedidos: {str(e)}'
            })
    
    return render(request, 'appConsultasTango/importFilePedidosAnularUY.html', {})

@login_required(login_url="/login/")
def ejecutar_anulacion_pedidos(request):
    """Vista para ejecutar la anulación de pedidos seleccionados"""
    nombre_db='LAKER_SA'
    settings.DATABASES['mi_db_2']['NAME'] = nombre_db
    
    if request.method == 'POST':
        try:
            # Obtener pedidos seleccionados del POST
            pedidos_seleccionados = request.POST.getlist('pedidos_seleccionados')
            anular_todos = request.POST.get('anular_todos') == 'true'
            
            if not pedidos_seleccionados:
                return render(request, 'appConsultasTango/importFilePedidosAnular.html', {
                    'mensaje_error': '❌ No se seleccionaron pedidos para anular'
                })
            
            # Procesar anulaciones
            anulados_exitosos = 0
            anulados_fallidos = 0
            
            for pedido_str in pedidos_seleccionados:
                try:
                    # Formato: "nro_pedido|talon_ped"
                    nro_pedido, talon_ped = pedido_str.split('|')
                    numero_pedido = ' 0000' + str(nro_pedido)[-9:]
                    
                    # Ejecutar anulación
                    anular_pedido(talon_ped, numero_pedido)
                    anulados_exitosos += 1
                except Exception as e:
                    print(f"Error anulando pedido {pedido_str}: {str(e)}")
                    anulados_fallidos += 1
            
            # Preparar mensaje de resultado
            if anulados_fallidos == 0:
                mensaje_success = f'✅ Se anularon exitosamente {anulados_exitosos} pedido(s)'
            else:
                mensaje_success = f'⚠️ Se anularon {anulados_exitosos} pedido(s). Fallaron {anulados_fallidos}.'
            
            return render(request, 'appConsultasTango/importFilePedidosAnular.html', {
                'mensaje_Success': mensaje_success
            })
            
        except Exception as e:
            print(f"Error en ejecutar_anulacion_pedidos: {str(e)}")
            return render(request, 'appConsultasTango/importFilePedidosAnular.html', {
                'mensaje_error': f'❌ Error al anular pedidos: {str(e)}'
            })
    
    return render(request, 'appConsultasTango/importFilePedidosAnular.html', {})

@login_required(login_url="/login/")
def validar_pedido_individual(request):
    """Vista para validar un pedido antes de anularlo (AR)"""
    nombre_db='LAKER_SA'
    settings.DATABASES['mi_db_2']['NAME'] = nombre_db
    
    if request.method == 'POST':
        nro_pedido = request.POST.get('nro_pedido', '').strip()
        talon_ped = request.POST.get('talon_ped', '').strip()
        
        if not nro_pedido or not talon_ped:
            return JsonResponse({
                'valido': False,
                'mensaje': '❌ Debes completar ambos campos (Nro. Pedido y Talón)'
            })
        
        try:
            # Validar usando la función existente
            resultado = validar_pedido(nro_pedido, talon_ped)
            
            if resultado > 0:
                return JsonResponse({
                    'valido': True,
                    'mensaje': f'✅ Pedido {nro_pedido} (Talón {talon_ped}) existe y está en estado PENDIENTE'
                })
            else:
                return JsonResponse({
                    'valido': False,
                    'mensaje': f'❌ El pedido {nro_pedido} (Talón {talon_ped}) no existe o no está en estado PENDIENTE'
                })
        except Exception as e:
            return JsonResponse({
                'valido': False,
                'mensaje': f'❌ Error al validar pedido: {str(e)}'
            })
    
    return JsonResponse({'valido': False, 'mensaje': 'Método no permitido'})

@login_required(login_url="/login/")
def validar_pedido_individual_uy(request):
    """Vista para validar un pedido antes de anularlo (UY)"""
    nombre_db='TASKY_SA'
    settings.DATABASES['mi_db_2']['NAME'] = nombre_db
    
    if request.method == 'POST':
        nro_pedido = request.POST.get('nro_pedido', '').strip()
        talon_ped = request.POST.get('talon_ped', '').strip()
        
        if not nro_pedido or not talon_ped:
            return JsonResponse({
                'valido': False,
                'mensaje': '❌ Debes completar ambos campos (Nro. Pedido y Talón)'
            })
        
        try:
            # Validar usando la función existente
            resultado = validar_pedido(nro_pedido, talon_ped)
            
            if resultado > 0:
                return JsonResponse({
                    'valido': True,
                    'mensaje': f'✅ Pedido {nro_pedido} (Talón {talon_ped}) existe y está en estado PENDIENTE'
                })
            else:
                return JsonResponse({
                    'valido': False,
                    'mensaje': f'❌ El pedido {nro_pedido} (Talón {talon_ped}) no existe o no está en estado PENDIENTE'
                })
        except Exception as e:
            return JsonResponse({
                'valido': False,
                'mensaje': f'❌ Error al validar pedido: {str(e)}'
            })
    
    return JsonResponse({'valido': False, 'mensaje': 'Método no permitido'})

def upload_file_AnularPedidos(path_filname):
    """Función legacy - mantener por compatibilidad"""
    excel_file =path_filname
    empexceldata = pd.read_excel(excel_file, engine='openpyxl')
    dbframe = empexceldata
    for df in dbframe.itertuples():
        numero_pedido = ' 0000' + str(df.NRO_PEDIDO)[-9:]
        talon_pedido = str(df.TALON_PED)
        anular_pedido(talon_pedido,numero_pedido)


@login_required(login_url="/login/")
def import_file_ubi(request):
    try:
        if request.method == 'POST' and request.FILES['excel_file']:
            pfile = request.FILES['excel_file']
            filesys =FileSystemStorage()
            uploadfilename = filesys.save(pfile.name,pfile) #Nombre del archivo
            extension = os.path.splitext(uploadfilename)
            if  not extension[1] == '.xlsx':
                error_extension = 'El formato del archivo debe ser de tipo .xlsx'
                os.remove(filesys.path(uploadfilename))
                return render(request,'appConsultasWMS/importFileUbi.html',{'mensaje_error':error_extension})

            uploaded_url = filesys.url(uploadfilename)  #Ruta donde se guardo el archivo
            uploaded_url = os.path.normpath(uploaded_url)
            # print("uploaded_url: " + uploaded_url)
            ruta_actual = os.path.join(os.getcwd()) #Ruta del proyecto
            # print('Ruta actual: ' + ruta_actual)
            path_filname = ruta_actual + uploaded_url
            wb = openpyxl.load_workbook(path_filname) #Abrimos el archivo
            worksheet = wb.active
            worksheet.cell(row=1, column=15).value = 'id_Ubicacion'
            excel_data = list()
            enc_data = list()
            mensaje_error = ''
            mensaje_Success = ''
            # iterando sobre las filas y obteniendo
            # valor de cada celda en la fila
            for row in worksheet.iter_rows():
                fila = row[0].row   #Numero de fila
                row_data = list()
                i = 1
                if worksheet.cell(row=fila, column=1).value is None: #Frena el loop al llegar al final de la lista
                    break
                
                for cell in row:
                    
                    # if cell.value == None:
                    #     break
                    if fila == 1:
                        enc_data.append(str(cell.value))
                    else:
                        if worksheet.cell(row=1, column=i).value == 'Cod_Ubicacion':
                        # print('Valor en contrado en la fila: ' + str(fila) + ' y columna: ' + str(i))
                            if fila > 1:
                                ubi = worksheet.cell(row=fila, column=i).value
                                # print(ubi)
                                Id_ubicacion= validar_ubicacion(ubi)
                                # print(Id_ubicacion)
                                if Id_ubicacion == 0:
                                    mensaje_error = 'Una o mas ubicaciones no existen en la base de datos'
                                    row_data.append('*' + str(cell.value) + '*')
                                    i += 1
                                    # print('row_data Id_ubicacion == 0: ')
                                    # print(row_data)
                                    continue
                                else:
                                    worksheet.cell(row=fila, column=15).value = Id_ubicacion

                        if cell.value == None:
                            row_data.append(str(''))
                        else:
                            row_data.append(str(cell.value))                        
                    i += 1

                excel_data.append(row_data)

            # print('path_filname: ' + path_filname)
            wb.save(path_filname)
            if not(mensaje_error):
                mensaje_Success = 'Se importo correctamente el archivo'
                upload_file_ubi(path_filname) #Ejecuta ejuste en base de datos

            # os.remove(filesys.path(uploadfilename))
            return render(request, 'appConsultasWMS/importFileUbi.html' ,{'enc_data':enc_data,'excel_data':excel_data,'mensaje_Success':mensaje_Success,'mensaje_error':mensaje_error})

        if request.method == 'POST' and request.SUBMIT['Procesar']:
            print('Hola Mundo >>>>')

    except Exception as identifier:            
        print(identifier)
    return render(request,'appConsultasWMS/importFileUbi.html',{})

def upload_file_ubi(path_filname):
    excel_file =path_filname
    empexceldata = pd.read_excel(excel_file, engine='openpyxl')
    dbframe = empexceldata

    for df in dbframe.itertuples():
        id_Ubi = str(df.id_Ubicacion)
        nom_ubi = df.Nombre_Ubicacion
        tipo = df.Tipo_Ubicacion
        estado_ubi = df.Estado_U

        if isnan(df.Rack):
            orden_rack = 0
        else:
            orden_rack = int(df.Rack)

        if isnan(df.Modulo):
            orden_modulo = 0
        else:
            orden_modulo = int(df.Modulo)
        
        if isnan(df.Altura):
            orden_altura = 0
        else:
            orden_altura = int(df.Altura)
        
            
        
        actualizar_ubicacion(id_Ubi,nom_ubi,tipo,estado_ubi,orden_rack,orden_modulo,orden_altura)

# ========================================
# CRUD para EB_sincArt_volumen
# ========================================

from .forms import EBSincArtVolumenForm
from .sql_volumen import (
    obtener_articulos_volumen, 
    obtener_articulo_volumen_por_codigo,
    actualizar_articulo_volumen,
    crear_articulo_volumen,
    eliminar_articulo_volumen,
    buscar_articulos_volumen,
    obtener_rubros_disponibles
)
from .excel_utils import generar_plantilla_excel, procesar_archivo_excel
from django.http import HttpResponse
import io

@login_required(login_url="/login/")
def eb_sinc_art_volumen_list(request):
    """Lista todos los artículos de volumen con funcionalidad de búsqueda y filtros"""
    # Configurar la base de datos
    nombre_db = 'LAKER_SA'
    settings.DATABASES['mi_db_2']['NAME'] = nombre_db
    
    search_term = request.GET.get('search', '')
    filtro_estado = request.GET.get('estado', '')
    filtro_rubro = request.GET.get('rubro', '')
    
    # Convertir filtro_estado a formato esperado por la función SQL
    estado_sql = None
    if filtro_estado == 'activo':
        estado_sql = 'activo'
    elif filtro_estado == 'inactivo':
        estado_sql = 'inactivo'
    
    if search_term:
        articulos = buscar_articulos_volumen(search_term, estado_sql, filtro_rubro or None)
    else:
        articulos = obtener_articulos_volumen(estado_sql, filtro_rubro or None)
    
    # Obtener rubros disponibles para el filtro
    rubros_disponibles = obtener_rubros_disponibles()
    
    # Paginación
    paginator = Paginator(articulos, 25)  # 25 artículos por página
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    context = {
        'page_obj': page_obj,
        'search_term': search_term,
        'filtro_estado': filtro_estado,
        'filtro_rubro': filtro_rubro,
        'rubros_disponibles': rubros_disponibles,
        'total_articulos': len(articulos)
    }
    
    return render(request, 'herramientas/eb_sinc_art_volumen/list.html', context)

@login_required(login_url="/login/")
def eb_sinc_art_volumen_detail(request, cod_articulo):
    """Muestra los detalles de un artículo específico"""
    # Configurar la base de datos
    nombre_db = 'LAKER_SA'
    settings.DATABASES['mi_db_2']['NAME'] = nombre_db
    
    articulo = obtener_articulo_volumen_por_codigo(cod_articulo)
    
    if not articulo:
        messages.error(request, 'Artículo no encontrado.')
        return redirect('herramientas:eb_sinc_art_volumen_list')
    
    return render(request, 'herramientas/eb_sinc_art_volumen/detail.html', {'articulo': articulo})

@login_required(login_url="/login/")
def eb_sinc_art_volumen_edit(request, cod_articulo):
    """Edita un artículo existente"""
    # Configurar la base de datos
    nombre_db = 'LAKER_SA'
    settings.DATABASES['mi_db_2']['NAME'] = nombre_db
    
    articulo = obtener_articulo_volumen_por_codigo(cod_articulo)
    
    if not articulo:
        messages.error(request, 'Artículo no encontrado.')
        return redirect('herramientas:eb_sinc_art_volumen_list')
    
    if request.method == 'POST':
        form = EBSincArtVolumenForm(request.POST, initial_data=articulo)
        if form.is_valid():
            try:
                # Ahora guardamos directamente en cm (no hay conversión)
                datos_actualizados = {
                    'alto_embalaje': form.cleaned_data['alto_embalaje'],
                    'ancho_embalaje': form.cleaned_data['ancho_embalaje'],
                    'largo_embalaje': form.cleaned_data['largo_embalaje'],
                    'alto_real': form.cleaned_data['alto_real'],
                    'ancho_real': form.cleaned_data['ancho_real'],
                    'largo_real': form.cleaned_data['largo_real'],
                    'peso_embalaje': form.cleaned_data['peso_embalaje'],
                    'peso_real': form.cleaned_data['peso_real']
                }
                
                actualizar_articulo_volumen(cod_articulo, datos_actualizados)
                messages.success(request, f'Artículo {cod_articulo} actualizado exitosamente.')
                return redirect('herramientas:eb_sinc_art_volumen_detail', cod_articulo=cod_articulo)
                
            except Exception as e:
                messages.error(request, f'Error al actualizar el artículo: {str(e)}')
    else:
        form = EBSincArtVolumenForm(initial_data=articulo)
    
    context = {
        'form': form,
        'articulo': articulo,
        'cod_articulo': cod_articulo
    }
    
    return render(request, 'herramientas/eb_sinc_art_volumen/edit.html', context)

@login_required(login_url="/login/")
def eb_sinc_art_volumen_delete(request, cod_articulo):
    """Elimina un artículo"""
    # Configurar la base de datos
    nombre_db = 'LAKER_SA'
    settings.DATABASES['mi_db_2']['NAME'] = nombre_db
    
    articulo = obtener_articulo_volumen_por_codigo(cod_articulo)
    
    if not articulo:
        messages.error(request, 'Artículo no encontrado.')
        return redirect('herramientas:eb_sinc_art_volumen_list')
    
    if request.method == 'POST':
        try:
            eliminar_articulo_volumen(cod_articulo)
            messages.success(request, f'Artículo {cod_articulo} eliminado exitosamente.')
            return redirect('herramientas:eb_sinc_art_volumen_list')
        except Exception as e:
            messages.error(request, f'Error al eliminar el artículo: {str(e)}')
    
    return render(request, 'herramientas/eb_sinc_art_volumen/delete.html', {'articulo': articulo})

@login_required(login_url="/login/")
def eb_sinc_art_volumen_descargar_plantilla(request):
    """Genera y descarga una plantilla Excel para carga masiva"""
    # Configurar la base de datos
    nombre_db = 'LAKER_SA'
    settings.DATABASES['mi_db_2']['NAME'] = nombre_db
    
    try:
        # Usar la función simplificada que sabemos que funciona
        from .excel_utils_simple import generar_plantilla_excel_simple
        wb = generar_plantilla_excel_simple()
        
        # Usar BytesIO para manejar el contenido en memoria
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        content = output.getvalue()
        
        # Crear respuesta HTTP con archivo Excel
        response = HttpResponse(
            content,
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename="Plantilla_Volumenes_Articulos.xlsx"'
        response['Content-Length'] = str(len(content))
        
        return response
        
    except Exception as e:
        messages.error(request, f'Error al generar la plantilla: {str(e)}')
        return redirect('herramientas:eb_sinc_art_volumen_list')

@login_required(login_url="/login/")
def eb_sinc_art_volumen_carga_masiva(request):
    """Maneja la carga masiva desde archivo Excel"""
    # Configurar la base de datos
    nombre_db = 'LAKER_SA'
    settings.DATABASES['mi_db_2']['NAME'] = nombre_db
    
    if request.method == 'POST':
        if 'archivo_excel' not in request.FILES:
            messages.error(request, 'No se seleccionó ningún archivo.')
            return redirect('herramientas:eb_sinc_art_volumen_list')
        
        archivo = request.FILES['archivo_excel']
        
        # Validar extensión
        if not archivo.name.endswith('.xlsx'):
            messages.error(request, 'El archivo debe tener extensión .xlsx')
            return redirect('herramientas:eb_sinc_art_volumen_list')
        
        try:
            errores, actualizaciones = procesar_archivo_excel(archivo)
            
            if errores:
                error_msg = f'Se procesaron {actualizaciones} registros correctamente. Errores encontrados:<br>'
                error_msg += '<br>'.join(errores[:10])  # Mostrar máximo 10 errores
                if len(errores) > 10:
                    error_msg += f'<br>... y {len(errores) - 10} errores más.'
                messages.warning(request, error_msg)
            else:
                messages.success(request, f'Carga masiva completada exitosamente. Se actualizaron {actualizaciones} artículos.')
            
        except Exception as e:
            messages.error(request, f'Error al procesar el archivo: {str(e)}')
    
    return redirect('herramientas:eb_sinc_art_volumen_list')

@login_required(login_url="/login/")
def eb_sinc_art_volumen_carga_masiva_form(request):
    """Muestra el formulario para carga masiva"""
    return render(request, 'herramientas/eb_sinc_art_volumen/carga_masiva.html')

def test_excel_download(request):
    """Vista de prueba simple para descargar Excel"""
    try:
        print("[TEST] Creando archivo Excel simple...")
        
        import openpyxl
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Test"
        
        ws['A1'] = "Prueba"
        ws['B1'] = "Excel"
        ws['A2'] = "Funciona"
        ws['B2'] = "OK"
        
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        response = HttpResponse(
            output.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename="test.xlsx"'
        
        print("[TEST] Archivo Excel de prueba creado correctamente")
        return response
        
    except Exception as e:
        print(f"[TEST ERROR] {str(e)}")
        return HttpResponse(f"Error: {str(e)}", content_type="text/plain")

def test_plantilla_step_by_step(request):
    """Prueba la generación de plantilla paso a paso"""
    # Configurar la base de datos
    nombre_db = 'LAKER_SA'
    settings.DATABASES['mi_db_2']['NAME'] = nombre_db
    
    try:
        print("[STEP 1] Configurando base de datos...")
        
        print("[STEP 2] Importando funciones...")
        from .sql_volumen import obtener_articulos_volumen
        
        print("[STEP 3] Obteniendo artículos...")
        articulos = obtener_articulos_volumen()
        print(f"[STEP 3] Se obtuvieron {len(articulos)} artículos")
        
        print("[STEP 4] Creando Excel básico con datos...")
        import openpyxl
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Volumenes"
        
        # Encabezados básicos
        ws['A1'] = "COD_ARTICULO"
        ws['B1'] = "DESCRIPCION"
        ws['C1'] = "RUBRO"
        
        # Agregar algunos datos (máximo 10 para prueba)
        for i, articulo in enumerate(articulos[:10], 2):
            ws[f'A{i}'] = articulo.get('COD_ARTICULO', '')
            ws[f'B{i}'] = articulo.get('DESCRIPCION', '')
            ws[f'C{i}'] = articulo.get('Rubro', '')
        
        print("[STEP 5] Generando respuesta...")
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        response = HttpResponse(
            output.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename="test_plantilla_simple.xlsx"'
        
        print("[STEP 5] ¡Éxito! Plantilla básica generada")
        return response
        
    except Exception as e:
        print(f"[STEP ERROR] Error en paso: {str(e)}")
        import traceback
        traceback.print_exc()
        return HttpResponse(f"Error en generación paso a paso: {str(e)}", content_type="text/plain")

def test_plantilla_simplificada(request):
    """Prueba con plantilla Excel simplificada"""
    # Configurar la base de datos
    nombre_db = 'LAKER_SA'
    settings.DATABASES['mi_db_2']['NAME'] = nombre_db
    
    try:
        print("[SIMPLE] Importando función simplificada...")
        from .excel_utils_simple import generar_plantilla_excel_simple
        
        print("[SIMPLE] Generando plantilla...")
        wb = generar_plantilla_excel_simple()
        
        print("[SIMPLE] Creando respuesta...")
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        response = HttpResponse(
            output.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename="Plantilla_Simplificada.xlsx"'
        
        print("[SIMPLE] ¡Éxito! Plantilla simplificada generada")
        return response
        
    except Exception as e:
        print(f"[SIMPLE ERROR] Error: {str(e)}")
        import traceback
        traceback.print_exc()
        return HttpResponse(f"Error en plantilla simplificada: {str(e)}", content_type="text/plain")

@login_required(login_url="/login/")
def gestion_sucursales_ecommerce(request):
    """Vista para gestionar las sucursales de ecommerce"""
    try:
        if request.method == 'POST':
            nro_sucursal = request.POST.get('nro_sucursal')
            if nro_sucursal:
                # Activar la sucursal seleccionada
                activar_sucursal_ecommerce(nro_sucursal)
                messages.success(request, f'Sucursal {nro_sucursal} activada exitosamente')
                return redirect('herramientas:gestion_sucursales_ecommerce')
        
        # Obtener todas las sucursales
        sucursales = obtener_sucursales_ecommerce()
        sucursal_activa = obtener_sucursal_activa()
        
        context = {
            'sucursales': sucursales,
            'sucursal_activa': sucursal_activa,
            'titulo': 'Gestión de Sucursales E-commerce'
        }
        
        return render(request, 'herramientas/gestion_sucursales_ecommerce.html', context)
        
    except Exception as e:
        messages.error(request, f'Error al cargar las sucursales: {str(e)}')
        return render(request, 'herramientas/gestion_sucursales_ecommerce.html', {
            'sucursales': [],
            'sucursal_activa': None,
            'titulo': 'Gestión de Sucursales E-commerce'
        })
