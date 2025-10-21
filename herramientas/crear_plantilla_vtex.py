import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import os

# Crear un nuevo libro de trabajo
libro = openpyxl.Workbook()
hoja = libro.active
hoja.title = 'Articulos VTEX'

# Estilos
header_font = Font(name='Arial', size=11, bold=True, color='000000')
header_fill = PatternFill(start_color='B4C7E7', end_color='B4C7E7', fill_type='solid')
header_alignment = Alignment(horizontal='center', vertical='center')

example_fill = PatternFill(start_color='E7E6E6', end_color='E7E6E6', fill_type='solid')
example_font = Font(name='Arial', size=10)

thin_border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

# Encabezados
encabezados = ['Codigo', 'Descripcion', 'DescripcionMetaTag']
for col_idx, encabezado in enumerate(encabezados, start=1):
    celda = hoja.cell(row=1, column=col_idx, value=encabezado)
    celda.font = header_font
    celda.fill = header_fill
    celda.alignment = header_alignment
    celda.border = thin_border

# Datos de ejemplo
ejemplos = [
    ['XV4SLL20A0101', 'CHELSEA PORTACOSMETICO', 'de la coleccion primavera-verano 2024/25 en cuotas sin interes en productos nacionales y envios a todo el pais.'],
    ['XV4SLL20A0109', 'CHELSEA PORTACOSMETICO', 'de la coleccion primavera-verano 2024/25 en cuotas sin interes en productos nacionales y envios a todo el pais.'],
    ['XV4SLL21A0101', 'PATTI COSMETIC', 'de la coleccion primavera-verano 2024/25 en cuotas sin interes en productos nacionales y envios a todo el pais.'],
    ['XV4SLL21B0558', 'PATTI FICHERO DOBLE', 'de la coleccion primavera-verano 2024/25 en cuotas sin interes en productos nacionales y envios a todo el pais.'],
    ['XV4SLL22B0301', 'CLARA FICH C SOLAPA Y CIERRE', 'de la coleccion primavera-verano 2024/25 en cuotas sin interes en productos nacionales y envios a todo el pais.']
]

for fila_idx, ejemplo in enumerate(ejemplos, start=2):
    for col_idx, valor in enumerate(ejemplo, start=1):
        celda = hoja.cell(row=fila_idx, column=col_idx, value=valor)
        celda.fill = example_fill
        celda.font = example_font
        celda.border = thin_border

# Ajustar ancho de columnas
hoja.column_dimensions['A'].width = 15
hoja.column_dimensions['B'].width = 30
hoja.column_dimensions['C'].width = 80

# Guardar archivo
script_dir = os.path.dirname(os.path.abspath(__file__))
project_root = os.path.dirname(script_dir)
media_dir = os.path.join(project_root, 'media')

if not os.path.exists(media_dir):
    os.makedirs(media_dir)

ruta_archivo = os.path.join(media_dir, 'AltaArtVtex.xlsx')
libro.save(ruta_archivo)

print('Plantilla .xlsx creada exitosamente en:', ruta_archivo)
print('Columnas: Codigo, Descripcion, DescripcionMetaTag')
print('5 filas de ejemplo incluidas')
