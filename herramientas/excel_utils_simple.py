import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from django.http import HttpResponse
from django.contrib import messages
from decimal import Decimal, InvalidOperation
import io
from .sql_volumen import obtener_articulos_volumen, actualizar_articulo_volumen

def generar_plantilla_excel_simple():
    """Genera una plantilla Excel completa y funcional para carga masiva de datos"""
    
    # Crear un nuevo workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Plantilla_Volumenes"
    
    # Definir estilos básicos
    try:
        from openpyxl.styles import Font, PatternFill, Alignment
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        required_fill = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")  
        editable_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
        center_align = Alignment(horizontal='center')
        estilos_disponibles = True
    except Exception as e:
        header_font = Font(bold=True)
        estilos_disponibles = False
    
    # Encabezados
    headers = [
        'COD_ARTICULO',
        'DESCRIPCION', 
        'RUBRO',
        'ALTO_EMBALAJE_CM',
        'ANCHO_EMBALAJE_CM', 
        'LARGO_EMBALAJE_CM',
        'ALTO_REAL_CM',
        'ANCHO_REAL_CM',
        'LARGO_REAL_CM',
        'PESO_EMBALAJE_G',
        'PESO_REAL_G'
    ]
    
    # Agregar encabezados con estilos
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.font = header_font
        if estilos_disponibles:
            cell.fill = header_fill
            cell.alignment = center_align
    
    
    # Obtener datos actuales para llenar la plantilla
    articulos = obtener_articulos_volumen()
    
    # Llenar datos (convertir mm a cm)
    for row_num, articulo in enumerate(articulos, 2):
        try:
            ws.cell(row=row_num, column=1, value=articulo['COD_ARTICULO'])
            ws.cell(row=row_num, column=2, value=articulo['DESCRIPCION'])
            ws.cell(row=row_num, column=3, value=articulo.get('Rubro', ''))
            
            # Convertir dimensiones de mm a cm (función robusta)
            def safe_mm_to_cm(mm_value):
                if mm_value is None or mm_value == 0:
                    return None
                try:
                    return round(float(mm_value) / 10.0, 2)
                except (ValueError, TypeError):
                    return None
            
            ws.cell(row=row_num, column=4, value=safe_mm_to_cm(articulo.get('altoEmbalaje')))
            ws.cell(row=row_num, column=5, value=safe_mm_to_cm(articulo.get('anchoEmbalaje')))
            ws.cell(row=row_num, column=6, value=safe_mm_to_cm(articulo.get('largoEmbalaje')))
            ws.cell(row=row_num, column=7, value=safe_mm_to_cm(articulo.get('altoReal')))
            ws.cell(row=row_num, column=8, value=safe_mm_to_cm(articulo.get('anchoReal')))
            ws.cell(row=row_num, column=9, value=safe_mm_to_cm(articulo.get('largoReal')))
            
            # Pesos se mantienen en gramos
            ws.cell(row=row_num, column=10, value=articulo.get('pesoEmbalaje'))
            ws.cell(row=row_num, column=11, value=articulo.get('pesoReal'))
            
            # Aplicar estilos de fondo si están disponibles
            if estilos_disponibles:
                # Campos no editables (naranja claro)
                for col_num in range(1, 4):
                    ws.cell(row=row_num, column=col_num).fill = required_fill
                # Campos editables (verde claro)  
                for col_num in range(4, 12):
                    ws.cell(row=row_num, column=col_num).fill = editable_fill
            
        except Exception as e:
            continue
    
    # Ajustar ancho de columnas
    column_widths = [15, 40, 20, 18, 18, 18, 15, 15, 15, 15, 15]
    for i, width in enumerate(column_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = width
    
    # Agregar hoja de instrucciones
    try:
        ws_instrucciones = wb.create_sheet("INSTRUCCIONES")
        
        instrucciones = [
            "INSTRUCCIONES PARA USO DE LA PLANTILLA",
            "",
            "1. CAMPOS OBLIGATORIOS (fondo naranja claro):",
            "   - COD_ARTICULO: Código único del artículo (NO MODIFICAR)",
            "   - DESCRIPCION: Descripción del artículo (NO MODIFICAR)", 
            "   - RUBRO: Rubro del artículo (NO MODIFICAR)",
            "",
            "2. CAMPOS EDITABLES (fondo verde claro):",
            "   - ALTO/ANCHO/LARGO_EMBALAJE_CM: Dimensiones en centímetros",
            "   - ALTO/ANCHO/LARGO_REAL_CM: Dimensiones reales en centímetros",
            "   - PESO_EMBALAJE_G/PESO_REAL_G: Pesos en gramos",
            "",
            "3. FORMATO:",
            "   - Use punto (.) como separador decimal: 12.5",
            "   - Las dimensiones están en centímetros (cm)",
            "   - Los pesos están en gramos (g)",
            "",
            "4. PROCESAMIENTO:",
            "   - Guarde el archivo manteniendo el formato Excel (.xlsx)",
            "   - Suba el archivo usando la función 'Carga Masiva'",
            "   - El sistema convertirá automáticamente cm a mm para almacenamiento"
        ]
        
        for row_num, texto in enumerate(instrucciones, 1):
            cell = ws_instrucciones.cell(row=row_num, column=1, value=texto)
            if row_num == 1:
                cell.font = Font(bold=True, size=14)
            elif texto.startswith(('1.', '2.', '3.', '4.')):
                cell.font = Font(bold=True)
        
        ws_instrucciones.column_dimensions['A'].width = 70
        
    except Exception as e:
        pass
    
    return wb