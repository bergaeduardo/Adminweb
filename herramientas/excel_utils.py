import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from django.http import HttpResponse
from django.contrib import messages
from decimal import Decimal, InvalidOperation
import io
from .sql_volumen import obtener_articulos_volumen, actualizar_articulo_volumen

def generar_plantilla_excel():
    """Genera una plantilla Excel para carga masiva de datos"""
    
    print("[DEBUG] Iniciando generación de plantilla Excel...")
    
    # Crear un nuevo workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Plantilla_Volumenes"
    
    print("[DEBUG] Creando estilos...")
    # Definir estilos (simplificados)
    try:
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)
        required_fill = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")
        editable_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
    except Exception as e:
        print(f"[DEBUG] Error creando estilos, usando básicos: {e}")
        header_fill = None
        header_font = Font(bold=True)
        required_fill = None
        editable_fill = None
    
    # Encabezados
    headers = [
        'COD_ARTICULO',
        'DESCRIPCION', 
        'RUBRO',
        'ALTO_EMBALAJE_CM',
        'ANCHO_EMBALAJE_CM', 
        'LARGO_EMBALAJE_CM',
        'ALTO_REAL_CM',
        'ANCHO_REAL_CM',
        'LARGO_REAL_CM',
        'PESO_EMBALAJE_G',
        'PESO_REAL_G'
    ]
    
    print("[DEBUG] Agregando encabezados...")
    # Agregar encabezados (sin border para simplificar)
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = header
        if header_fill:
            cell.fill = header_fill
        cell.font = header_font
    
    print("[DEBUG] Obteniendo datos de artículos...")
    # Obtener datos actuales para llenar la plantilla (limitar a 500 registros para evitar timeout)
    try:
        articulos = obtener_articulos_volumen()
        print(f"[DEBUG] Obtenidos {len(articulos)} artículos, procesando máximo 500...")
        articulos = articulos[:500]  # Limitar registros
    except Exception as e:
        print(f"[DEBUG] Error obteniendo artículos: {e}")
        articulos = []
    
    print("[DEBUG] Procesando datos de artículos...")
    # Llenar datos (convertir mm a cm)
    for row_num, articulo in enumerate(articulos, 2):
        try:
            ws.cell(row=row_num, column=1, value=articulo['COD_ARTICULO'])
            ws.cell(row=row_num, column=2, value=articulo['DESCRIPCION'])
            ws.cell(row=row_num, column=3, value=articulo.get('Rubro', ''))
            
            # Convertir dimensiones de mm a cm (función simplificada)
            def safe_mm_to_cm(valor):
                try:
                    if valor is None or valor == 0:
                        return None
                    return round(float(valor) / 10, 2)
                except:
                    return None
            
            ws.cell(row=row_num, column=4, value=safe_mm_to_cm(articulo.get('altoEmbalaje')))
            ws.cell(row=row_num, column=5, value=safe_mm_to_cm(articulo.get('anchoEmbalaje')))
            ws.cell(row=row_num, column=6, value=safe_mm_to_cm(articulo.get('largoEmbalaje')))
            ws.cell(row=row_num, column=7, value=safe_mm_to_cm(articulo.get('altoReal')))
            ws.cell(row=row_num, column=8, value=safe_mm_to_cm(articulo.get('anchoReal')))
            ws.cell(row=row_num, column=9, value=safe_mm_to_cm(articulo.get('largoReal')))
            
            # Pesos se mantienen en gramos
            ws.cell(row=row_num, column=10, value=articulo.get('pesoEmbalaje'))
            ws.cell(row=row_num, column=11, value=articulo.get('pesoReal'))
            
            # Aplicar colores de fondo solo si están disponibles
            if required_fill:
                for col_num in range(1, 4):  # Campos no editables
                    ws.cell(row=row_num, column=col_num).fill = required_fill
            if editable_fill:
                for col_num in range(4, 12):  # Campos editables
                    ws.cell(row=row_num, column=col_num).fill = editable_fill
                    
        except Exception as e:
            print(f"[DEBUG] Error procesando artículo {articulo.get('COD_ARTICULO', 'UNKNOWN')}: {e}")
            continue
    
    print("[DEBUG] Configurando columnas...")
    # Ajustar ancho de columnas
    try:
        column_widths = [15, 40, 20, 18, 18, 18, 15, 15, 15, 15, 15]
        for i, width in enumerate(column_widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = width
    except Exception as e:
        print(f"[DEBUG] Error configurando anchos: {e}")
    
    print("[DEBUG] Creando hoja de instrucciones...")
    # Agregar hoja de instrucciones (simplificada)
    try:
        ws_instrucciones = wb.create_sheet("INSTRUCCIONES")
        
        instrucciones = [
            "INSTRUCCIONES PARA USO DE LA PLANTILLA",
            "",
            "1. NO modificar: COD_ARTICULO, DESCRIPCION, RUBRO",
            "2. Modificar solo: dimensiones (en cm) y pesos (en g)",
            "3. Usar punto (.) como separador decimal: 12.5",
            "4. Guardar como .xlsx y subir con 'Carga Masiva'"
        ]
        
        for row_num, texto in enumerate(instrucciones, 1):
            ws_instrucciones.cell(row=row_num, column=1, value=texto)
        
        ws_instrucciones.column_dimensions['A'].width = 60
    except Exception as e:
        print(f"[DEBUG] Error creando instrucciones: {e}")
    
    print("[DEBUG] Plantilla Excel completada exitosamente")
    return wb

def convertir_mm_a_cm(valor_mm):
    """Convierte milímetros a centímetros"""
    if valor_mm is None or valor_mm == 0:
        return None
    try:
        return round(float(valor_mm) / 10, 2)
    except (ValueError, TypeError):
        return None

def convertir_cm_a_mm(valor_cm):
    """Convierte centímetros a milímetros"""
    if valor_cm is None or valor_cm == '' or valor_cm == 0:
        return None
    try:
        return round(float(valor_cm) * 10)
    except (ValueError, TypeError):
        return None

def procesar_archivo_excel(archivo_excel):
    """Procesa el archivo Excel y actualiza los datos"""
    
    errores = []
    actualizaciones_exitosas = 0
    
    try:
        # Leer el archivo Excel
        wb = openpyxl.load_workbook(archivo_excel)
        ws = wb.active
        
        # Validar encabezados
        encabezados_esperados = [
            'COD_ARTICULO', 'DESCRIPCION', 'RUBRO',
            'ALTO_EMBALAJE_CM', 'ANCHO_EMBALAJE_CM', 'LARGO_EMBALAJE_CM',
            'ALTO_REAL_CM', 'ANCHO_REAL_CM', 'LARGO_REAL_CM',
            'PESO_EMBALAJE_G', 'PESO_REAL_G'
        ]
        
        encabezados_archivo = []
        for col in range(1, 12):
            valor = ws.cell(row=1, column=col).value
            if valor:
                encabezados_archivo.append(str(valor).strip())
        
        if encabezados_archivo != encabezados_esperados:
            errores.append("Los encabezados del archivo no coinciden con la plantilla esperada")
            return errores, 0
        
        # Procesar cada fila
        for row_num in range(2, ws.max_row + 1):
            try:
                cod_articulo = ws.cell(row=row_num, column=1).value
                
                if not cod_articulo:
                    continue  # Saltar filas vacías
                
                cod_articulo = str(cod_articulo).strip()
                
                # Leer dimensiones (convertir de cm a mm)
                alto_embalaje = convertir_cm_a_mm(ws.cell(row=row_num, column=4).value)
                ancho_embalaje = convertir_cm_a_mm(ws.cell(row=row_num, column=5).value)
                largo_embalaje = convertir_cm_a_mm(ws.cell(row=row_num, column=6).value)
                alto_real = convertir_cm_a_mm(ws.cell(row=row_num, column=7).value)
                ancho_real = convertir_cm_a_mm(ws.cell(row=row_num, column=8).value)
                largo_real = convertir_cm_a_mm(ws.cell(row=row_num, column=9).value)
                
                # Leer pesos (mantener en gramos)
                peso_embalaje = validar_numero_entero(ws.cell(row=row_num, column=10).value)
                peso_real = validar_numero_entero(ws.cell(row=row_num, column=11).value)
                
                # Preparar datos para actualización
                datos_actualizados = {
                    'alto_embalaje': alto_embalaje,
                    'ancho_embalaje': ancho_embalaje,
                    'largo_embalaje': largo_embalaje,
                    'alto_real': alto_real,
                    'ancho_real': ancho_real,
                    'largo_real': largo_real,
                    'peso_embalaje': peso_embalaje,
                    'peso_real': peso_real
                }
                
                # Actualizar en base de datos
                actualizar_articulo_volumen(cod_articulo, datos_actualizados)
                actualizaciones_exitosas += 1
                
            except Exception as e:
                errores.append(f"Fila {row_num}: Error al procesar - {str(e)}")
                continue
        
    except Exception as e:
        errores.append(f"Error al leer el archivo: {str(e)}")
    
    return errores, actualizaciones_exitosas

def validar_numero_entero(valor):
    """Valida y convierte un valor a entero"""
    if valor is None or valor == '':
        return None
    try:
        return int(float(valor))
    except (ValueError, TypeError):
        return None

def validar_numero_decimal(valor):
    """Valida y convierte un valor a decimal"""
    if valor is None or valor == '':
        return None
    try:
        return float(valor)
    except (ValueError, TypeError):
        return None