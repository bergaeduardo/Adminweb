# -*- coding: utf-8 -*-
from __future__ import unicode_literals

import pprint
from tkinter.tix import CELL, COLUMN
from django import template
from django.contrib.auth.decorators import login_required
from django.http import HttpResponse, HttpResponseRedirect
from django.template import loader
from django.urls import reverse
from django.shortcuts import render, redirect
from numpy import int64, isnan
from consultasTango.models import StockCentral,SjStockDisponibleEcommerce
from consultasWMS.models import Ubicacion
from consultasLakersBis.models import Direccionario, SofStockLakers
from django.views.generic.list import ListView
from apps.home.vistas.settingsUrls import *
from apps.home.SQL.Sql_WMS import validar_ubicacion,actualizar_ubicacion
from apps.home.SQL.Sql_Tango import validar_pedido,cerrar_pedido,validar_articulo,borrar_contTabla,cargar_articulo
# from apps.static.Scripts.getData_Trello import reporte_trello
import json
from consultasTango.filters import *
from consultasLakersBis.filters import filtroCanal,filtroTipoLocal,filtroGrupoEmpresario,DireccionarioFilter
from django.contrib import messages
import openpyxl
import pandas as pd



from django.shortcuts import render, redirect 
from django.conf import settings
from django.core.files.storage import FileSystemStorage
from consultasLakersBis.models import SucursalesLakers
from consultasLakersBis.forms import sucursalesform
import os
import subprocess
from django.templatetags.static import static

@login_required(login_url="/login/")
def runscript(request):
    # script_path = 'ruta_al_script/script.py'
    # # Ejecutar el script utilizando subprocess
    # output = subprocess.check_output(['python', script_path])
    return render(request, 'appConsultasTango/runScript.html')


def runscriptResult(request):
    mensaje_success=''
    output=''
    ruta_actual = r'C:\Users\eduardo.berga\Desktop\Proyectos\Lakers_Lab\Adminweb\apps\static\Scripts'
    archivo = r'\miTareaPy.py'
    # script_path = static(r'Scripts\Imprimir_Etiquetas_Andreani2.py')
    # script_path= ruta_actual + '\Scripts\hora.py'
    # script_path = r'C:\Users\eduardo.berga\Desktop\Proyectos\Lakers_Lab\Adminweb\apps\static\Scripts\hora.py'
    script_path = ruta_actual + archivo
    print('Ruta actual: ' + script_path)
    # # Ejecutar el script utilizando subprocess
    resultado = subprocess.run(['python', script_path], capture_output=True, text=True)
    salida = resultado.returncode
    if salida == 0:
        mensaje_success = "La ejecucion fue un exito"
        output = 'El resultado del script es la hora actual: ' + resultado.stdout
    print(resultado)
    # output = subprocess.check_output(['python', script_path])
    # output = "Hola Mundo"
    return render(request, 'appConsultasTango/runScriptResult.html', {'mensaje_success': mensaje_success,'output': output})


@login_required(login_url="/login/")
def editarSucursal(request,id):
    suc = SucursalesLakers.objects.get(nro_sucursal=id)
    sucForm = sucursalesform(request.POST or None, request.FILES or None, instance=suc)
    Disabled='disabled'
    # print(id)
    # print(sucForm.errors)
    if sucForm.is_valid() and request.POST:
        suc = sucForm.save(commit=False)
        suc.save()
        messages.success(request, 'OK')        
        infForm = sucForm.cleaned_data
        # print(infForm)
        return redirect('/../Extras/direccionario')

    return  render(request,'appConsultasTango/editarSucursal.html',{'formulario':sucForm,'Disabled':Disabled})

@login_required(login_url="/login/")
def registraSucursal(request):
    if request.method=='POST':
        formulario=sucursalesform(request.POST)
        if formulario.is_valid():
            sucursal = formulario.save(commit=False)
            sucursal.save()
            messages.success(request, 'OK')
            infForm = formulario.cleaned_data
            # print(infForm)
            return redirect('/../Extras/direccionario')
        
    else: 
        formulario=sucursalesform()

    return  render(request,'appConsultasTango/registraSucursal.html',{'formulario':formulario})

@login_required(login_url="/login/")
def import_file_etiquetas(request):
    nombre_db='LAKER_SA'
    settings.DATABASES['mi_db_2']['NAME'] = nombre_db
    print('Cambiando base de datos a ' + nombre_db)
    print('import_file_etiquetas')
    try:
        if request.method == 'POST' and request.FILES['excel_file']:
            pfile = request.FILES['excel_file']
            filesys =FileSystemStorage()
            # Obtener el nombre del archivo sin espacios en blanco
            filename = pfile.name.replace(' ', '')

            # Guardar el archivo con el nombre sin espacios en blanco
            uploadfilename = filesys.save(filename, pfile)
            extension = os.path.splitext(uploadfilename)
            if  not extension[1] == '.xlsx':
                error_extension = 'El formato del archivo debe ser de tipo .xlsx'
                os.remove(filesys.path(uploadfilename))
                return render(request,'appConsultasTango/importFileEtiquetas.html',{'mensaje_error':error_extension})

            uploaded_url = filesys.url(uploadfilename)  #Ruta donde se guardo el archivo
            uploaded_url = os.path.normpath(uploaded_url)
            # print("uploaded_url: " + uploaded_url)
            ruta_actual = os.path.join(os.getcwd()) #Ruta del proyecto
            # print('Ruta actual: ' + ruta_actual)
            path_filname = ruta_actual + uploaded_url
            wb = openpyxl.load_workbook(path_filname) #Abrimos el archivo
            worksheet = wb.active
            excel_data = list()
            enc_data = list()
            mensaje_error = ''
            mensaje_Success = ''
            art_valido = ''
            # iterando sobre las filas y obteniendo
            # valor de cada celda en la fila
            for row in worksheet.iter_rows():
                fila = row[0].row   #Numero de fila
                row_data = list()
                talon_pedido = 0
                i = 1
                if worksheet.cell(row=fila, column=1).value is None: #Frena el loop al llegar al final de la lista
                    break
                
                for cell in row:
                    if fila == 1:
                        enc_data.append(str(cell.value))
                    else:
                        if worksheet.cell(row=1, column=i).value == 'ARTICULO':
                            if fila > 1:
                                articulo = worksheet.cell(row=fila, column=i).value
                                art_valido= validar_articulo(articulo)
                                print('art_valido: ' + str(art_valido))
                                if art_valido == 'ERROR':
                                    mensaje_error = 'Hay articulos que no existen en SJ_ETIQUETAS_FINAL'
                                    row_data.append('*' + str(cell.value) + '*')
                                    i += 1
                                    continue
                        elif worksheet.cell(row=1, column=i).value == 'DESCRIPCION':
                            worksheet.cell(row=fila, column=i).value = art_valido

                        if cell.value == None:
                            row_data.append(str(''))
                        else:
                            row_data.append(str(cell.value))                        
                    i += 1

                excel_data.append(row_data)

            # print('path_filname: ' + path_filname)
            wb.save(path_filname)
            if not(mensaje_error):
                upload_file_artEtiquetas(path_filname) #Carga articulos en base de datos
                mensaje_Success = 'Articulos cargados correctamente'
                os.remove(filesys.path(uploadfilename))
                
            else:
                os.remove(filesys.path(uploadfilename))

            
            return render(request, 'appConsultasTango/importFileEtiquetas.html' ,{'enc_data':enc_data,'excel_data':excel_data,'mensaje_Success':mensaje_Success,'mensaje_error':mensaje_error})


    except Exception as identifier:            
        print(identifier)
    return render(request,'appConsultasTango/importFileEtiquetas.html',{})

def upload_file_artEtiquetas(path_filname):
    excel_file =path_filname
    empexceldata = pd.read_excel(excel_file)
    dbframe = empexceldata
    borrar_contTabla('SJ_T_ETIQUETAS_FINAL')
    
    for df in dbframe.itertuples():
        
        articulo = df.ARTICULO
        descripcion = str(df.DESCRIPCION)
        cargar_articulo(articulo,descripcion)

@login_required(login_url="/login/")
def import_file_cierrePedidos(request):
    nombre_db='LAKER_SA'
    settings.DATABASES['mi_db_2']['NAME'] = nombre_db
    print('Cambiando base de datos a ' + nombre_db)
    try:
        if request.method == 'POST' and request.FILES['excel_file']:
            pfile = request.FILES['excel_file']
            filesys =FileSystemStorage()
            # Obtener el nombre del archivo sin espacios en blanco
            filename = pfile.name.replace(' ', '')

            # Guardar el archivo con el nombre sin espacios en blanco
            uploadfilename = filesys.save(filename, pfile)
            extension = os.path.splitext(uploadfilename)
            if  not extension[1] == '.xlsx':
                error_extension = 'El formato del archivo debe ser de tipo .xlsx'
                os.remove(filesys.path(uploadfilename))
                return render(request,'appConsultasTango/importFilePedidosCierre.html',{'mensaje_error':error_extension})

            uploaded_url = filesys.url(uploadfilename)  #Ruta donde se guardo el archivo
            uploaded_url = os.path.normpath(uploaded_url)
            # print("uploaded_url: " + uploaded_url)
            ruta_actual = os.path.join(os.getcwd()) #Ruta del proyecto
            # print('Ruta actual: ' + ruta_actual)
            path_filname = ruta_actual + uploaded_url
            wb = openpyxl.load_workbook(path_filname) #Abrimos el archivo
            worksheet = wb.active
            excel_data = list()
            enc_data = list()
            mensaje_error = ''
            mensaje_Success = ''
            # iterando sobre las filas y obteniendo
            # valor de cada celda en la fila
            for row in worksheet.iter_rows():
                fila = row[0].row   #Numero de fila
                row_data = list()
                talon_pedido = 0
                i = 1
                if worksheet.cell(row=fila, column=1).value is None: #Frena el loop al llegar al final de la lista
                    break
                
                for cell in row:
                    if fila == 1:
                        enc_data.append(str(cell.value))
                    else:
                        if worksheet.cell(row=1, column=i).value == 'NRO_PEDIDO':
                            if fila > 1:
                                pedido = worksheet.cell(row=fila, column=i).value
                                if worksheet.cell(row=1, column=6).value == 'TALON_PED':
                                    talon_pedido = worksheet.cell(row=fila, column=6).value

                                ped_valido= validar_pedido(pedido,str(talon_pedido))
                                # print('ped_valido: ' + str(ped_valido))
                                if ped_valido == 0:
                                    mensaje_error = 'Los Pedidos NO EXISTEN  o NO ESTAN PENDIENTES'
                                    row_data.append('*' + str(cell.value) + '*')
                                    i += 1
                                    continue
                        if cell.value == None:
                            row_data.append(str(''))
                        else:
                            row_data.append(str(cell.value))                        
                    i += 1

                excel_data.append(row_data)

            # print('path_filname: ' + path_filname)
            wb.save(path_filname)
            if not(mensaje_error):
                upload_file_CierrePedidos(path_filname) #Ejecuta ejuste en base de datos
                mensaje_Success = 'Articulos cargados correctamente'
                os.remove(filesys.path(uploadfilename))
                
            else:
                os.remove(filesys.path(uploadfilename))

            
            return render(request, 'appConsultasTango/importFilePedidosCierre.html' ,{'enc_data':enc_data,'excel_data':excel_data,'mensaje_Success':mensaje_Success,'mensaje_error':mensaje_error})


    except Exception as identifier:            
        print(identifier)
    return render(request,'appConsultasTango/importFilePedidosCierre.html',{})

@login_required(login_url="/login/")
def import_file_cierrePedidosUY(request):
    nombre_db='TASKY_SA'
    settings.DATABASES['mi_db_2']['NAME'] = nombre_db
    print('Cambiando base de datos a ' + nombre_db)
    try:
        if request.method == 'POST' and request.FILES['excel_file']:
            pfile = request.FILES['excel_file']
            filesys =FileSystemStorage()
            # Obtener el nombre del archivo sin espacios en blanco
            filename = pfile.name.replace(' ', '')

            # Guardar el archivo con el nombre sin espacios en blanco
            uploadfilename = filesys.save(filename, pfile)
            extension = os.path.splitext(uploadfilename)
            if  not extension[1] == '.xlsx':
                error_extension = 'El formato del archivo debe ser de tipo .xlsx'
                os.remove(filesys.path(uploadfilename))
                return render(request,'appConsultasTango/importFilePedidosCierreUY.html',{'mensaje_error':error_extension})

            uploaded_url = filesys.url(uploadfilename)  #Ruta donde se guardo el archivo
            uploaded_url = os.path.normpath(uploaded_url)
            # print("uploaded_url: " + uploaded_url)
            ruta_actual = os.path.join(os.getcwd()) #Ruta del proyecto
            # print('Ruta actual: ' + ruta_actual)
            path_filname = ruta_actual + uploaded_url
            wb = openpyxl.load_workbook(path_filname) #Abrimos el archivo
            worksheet = wb.active
            excel_data = list()
            enc_data = list()
            mensaje_error = ''
            mensaje_Success = ''
            # iterando sobre las filas y obteniendo
            # valor de cada celda en la fila
            for row in worksheet.iter_rows():
                fila = row[0].row   #Numero de fila
                row_data = list()
                talon_pedido = 0
                i = 1
                if worksheet.cell(row=fila, column=1).value is None: #Frena el loop al llegar al final de la lista
                    break
                
                for cell in row:
                    if fila == 1:
                        enc_data.append(str(cell.value))
                    else:
                        if worksheet.cell(row=1, column=i).value == 'NRO_PEDIDO':
                            if fila > 1:
                                pedido = worksheet.cell(row=fila, column=i).value
                                if worksheet.cell(row=1, column=6).value == 'TALON_PED':
                                    talon_pedido = worksheet.cell(row=fila, column=6).value

                                ped_valido= validar_pedido(pedido,str(talon_pedido))
                                # print('ped_valido: ' + str(ped_valido))
                                if ped_valido == 0:
                                    mensaje_error = 'Los Pedidos NO EXISTEN  o NO ESTAN PENDIENTES'
                                    row_data.append('*' + str(cell.value) + '*')
                                    i += 1
                                    continue
                        if cell.value == None:
                            row_data.append(str(''))
                        else:
                            row_data.append(str(cell.value))                        
                    i += 1

                excel_data.append(row_data)

            # print('path_filname: ' + path_filname)
            wb.save(path_filname)
            if not(mensaje_error):
                upload_file_CierrePedidos(path_filname) #Ejecuta ejuste en base de datos
                mensaje_Success = 'Pedidos anulados correctamente'
                os.remove(filesys.path(uploadfilename))
                
            else:
                os.remove(filesys.path(uploadfilename))

            
            return render(request, 'appConsultasTango/importFilePedidosCierreUY.html' ,{'enc_data':enc_data,'excel_data':excel_data,'mensaje_Success':mensaje_Success,'mensaje_error':mensaje_error})


    except Exception as identifier:            
        print(identifier)
    return render(request,'appConsultasTango/importFilePedidosCierreUY.html',{})

def upload_file_CierrePedidos(path_filname):
    excel_file =path_filname
    empexceldata = pd.read_excel(excel_file)
    dbframe = empexceldata
    
    for df in dbframe.itertuples():
        
        numero_pedido = df.NRO_PEDIDO
        talon_pedido = str(df.TALON_PED)

        if type(numero_pedido) == int:
            pedido = ' 0000' + str(numero_pedido)
        else:
            pedido = numero_pedido
            print('Dato sin concatenar')
        
        cerrar_pedido(talon_pedido,pedido)


@login_required(login_url="/login/")
def import_file_ubi(request):
    try:
        if request.method == 'POST' and request.FILES['excel_file']:
            pfile = request.FILES['excel_file']
            filesys =FileSystemStorage()
            uploadfilename = filesys.save(pfile.name,pfile) #Nombre del archivo
            extension = os.path.splitext(uploadfilename)
            if  not extension[1] == '.xlsx':
                error_extension = 'El formato del archivo debe ser de tipo .xlsx'
                os.remove(filesys.path(uploadfilename))
                return render(request,'appConsultasWMS/importFileUbi.html',{'mensaje_error':error_extension})

            uploaded_url = filesys.url(uploadfilename)  #Ruta donde se guardo el archivo
            uploaded_url = os.path.normpath(uploaded_url)
            # print("uploaded_url: " + uploaded_url)
            ruta_actual = os.path.join(os.getcwd()) #Ruta del proyecto
            # print('Ruta actual: ' + ruta_actual)
            path_filname = ruta_actual + uploaded_url
            wb = openpyxl.load_workbook(path_filname) #Abrimos el archivo
            worksheet = wb.active
            worksheet.cell(row=1, column=15).value = 'id_Ubicacion'
            excel_data = list()
            enc_data = list()
            mensaje_error = ''
            mensaje_Success = ''
            # iterando sobre las filas y obteniendo
            # valor de cada celda en la fila
            for row in worksheet.iter_rows():
                fila = row[0].row   #Numero de fila
                row_data = list()
                i = 1
                if worksheet.cell(row=fila, column=1).value is None: #Frena el loop al llegar al final de la lista
                    break
                
                for cell in row:
                    
                    # if cell.value == None:
                    #     break
                    if fila == 1:
                        enc_data.append(str(cell.value))
                    else:
                        if worksheet.cell(row=1, column=i).value == 'Cod_Ubicacion':
                        # print('Valor en contrado en la fila: ' + str(fila) + ' y columna: ' + str(i))
                            if fila > 1:
                                ubi = worksheet.cell(row=fila, column=i).value
                                # print(ubi)
                                Id_ubicacion= validar_ubicacion(ubi)
                                # print(Id_ubicacion)
                                if Id_ubicacion == 0:
                                    mensaje_error = 'Una o mas ubicaciones no existen en la base de datos'
                                    row_data.append('*' + str(cell.value) + '*')
                                    i += 1
                                    # print('row_data Id_ubicacion == 0: ')
                                    # print(row_data)
                                    continue
                                else:
                                    worksheet.cell(row=fila, column=15).value = Id_ubicacion

                        if cell.value == None:
                            row_data.append(str(''))
                        else:
                            row_data.append(str(cell.value))                        
                    i += 1

                excel_data.append(row_data)

            # print('path_filname: ' + path_filname)
            wb.save(path_filname)
            if not(mensaje_error):
                mensaje_Success = 'Se importo correctamente el archivo'
                upload_file_ubi(path_filname) #Ejecuta ejuste en base de datos

            # os.remove(filesys.path(uploadfilename))
            return render(request, 'appConsultasWMS/importFileUbi.html' ,{'enc_data':enc_data,'excel_data':excel_data,'mensaje_Success':mensaje_Success,'mensaje_error':mensaje_error})

        if request.method == 'POST' and request.SUBMIT['Procesar']:
            print('Hola Mundo >>>>')

    except Exception as identifier:            
        print(identifier)
    return render(request,'appConsultasWMS/importFileUbi.html',{})

def upload_file_ubi(path_filname):
    excel_file =path_filname
    empexceldata = pd.read_excel(excel_file)
    dbframe = empexceldata

    for df in dbframe.itertuples():
        id_Ubi = str(df.id_Ubicacion)
        nom_ubi = df.Nombre_Ubicacion
        tipo = df.Tipo_Ubicacion
        estado_ubi = df.Estado_U

        if isnan(df.Rack):
            orden_rack = 0
        else:
            orden_rack = int(df.Rack)

        if isnan(df.Modulo):
            orden_modulo = 0
        else:
            orden_modulo = int(df.Modulo)
        
        if isnan(df.Altura):
            orden_altura = 0
        else:
            orden_altura = int(df.Altura)
        
            
        
        actualizar_ubicacion(id_Ubi,nom_ubi,tipo,estado_ubi,orden_rack,orden_modulo,orden_altura)

@login_required(login_url="/login/") # DESABILITADO
def direccionario(request):
    Nombre='Pedidos'
    dir_iframe = DIR_EXTRAS['direccionario']
    return render(request,'home/direccionario.html',{'dir_iframe':dir_iframe,'Nombre':Nombre})

@login_required(login_url="/login/")
def agenda(request):    # <<<----- Direccionario Cards -->
    Nombre='Direccionario'
    datos = Direccionario.objects.all()
    canal = filtroCanal()
    tipo_local = filtroTipoLocal()
    grupo = filtroGrupoEmpresario()

    return render(request,'appConsultasTango/direccionario2.html',{'datos': datos,'canal':canal,'tipo_local':tipo_local,'grupo':grupo,'Nombre':Nombre})

@login_required(login_url="/login/")
def DireccionarioTabla(request):    # <<<----- Direccionario Tabla -->
    Nombre='Direccionario'
    stock = Direccionario.objects.all()
    myFilter = DireccionarioFilter(request.GET, queryset=stock)
    if request.GET:
        datos = myFilter
    else:
        datos = Direccionario.objects.all()

    return render(request,'appConsultasTango/direccionarioTabla.html',{'myFilter':myFilter,'datos':datos,'Nombre':Nombre})

# @login_required(login_url="/login/")
# def reporTrello(request):
#     miReporte_json = reporte_trello()
# #     print(type(miReporte_json))

# # # convertir la variable miReporte_json del formato <class 'dict'> a json
#     miReporte_json = json.dumps(miReporte_json)
#     print(type(miReporte_json))
#     # print(miReporte_json)

#     return render(request, 'home/trello-activity-report.html', {'data': miReporte_json})

@login_required(login_url="/login/")
def reporTrello(request):
    Nombre=''
    dir_iframe = DIR_EXTRAS['reporteTrello']
    return render(request,'home/PlantillaReportes.html',{'dir_iframe':dir_iframe,'Nombre':Nombre})

@login_required(login_url="/login/")
def internos(request):
    Nombre=''
    dir_iframe = DIR_EXTRAS['internos']
    return render(request,'home/PlantillaReportes.html',{'dir_iframe':dir_iframe,'Nombre':Nombre})
